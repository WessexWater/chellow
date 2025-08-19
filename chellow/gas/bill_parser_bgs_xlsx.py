import csv
from collections import defaultdict
from datetime import datetime as Datetime
from decimal import Decimal, InvalidOperation
from enum import Enum, auto
from io import BytesIO

from dateutil.relativedelta import relativedelta

from openpyxl import load_workbook

from werkzeug.exceptions import BadRequest

from chellow.utils import hh_max, hh_min, to_ct, to_utc


class Title(Enum):
    CUSTOMER = auto()
    PRODUCT = auto()
    BROKER_NAME = auto()
    ACCOUNT = auto()
    MPRN = auto()
    BILL = auto()
    SUPPLY_ADDRESS = auto()
    BILL_DATE = auto()
    BILLING_PERIOD = auto()
    CHARGE_TYPE = auto()
    CHARGE_PERIOD_FROM = auto()
    CHARGE_PERIOD_END = auto()
    QUANTITY = auto()
    QUNIT = auto()
    CHARGE = auto()
    CUNIT = auto()
    TOTAL = auto()


COLUMNS = {
    Title.CUSTOMER: ["Customer"],
    Title.ACCOUNT: ["Account"],
    Title.MPRN: ["MPRN"],
    Title.BILL: ["Bill"],
    Title.BILL_DATE: ["Bill Date", "BillDate"],
    Title.BILLING_PERIOD: ["Billing Period", "BillingPeriod"],
    Title.CHARGE_TYPE: ["Charge Type", "ChargeType"],
    Title.CHARGE_PERIOD_FROM: ["Charge Period From", "ChargePeriodFrom"],
    Title.CHARGE_PERIOD_END: ["Charge Period End", "ChargePeriodEnd"],
    Title.QUANTITY: ["Quantity"],
    Title.QUNIT: ["QUnit"],
    Title.CHARGE: ["Charge"],
    Title.CUNIT: ["CUnit"],
    Title.TOTAL: ["Total"],
}


def make_column_map(title_row):
    titles = [cell.value for cell in title_row]
    column_map = {}
    for title, title_names in COLUMNS.items():
        idx = None
        for title_name in title_names:
            try:
                idx = titles.index(title_name)
            except ValueError:
                pass
        if idx is None:
            raise BadRequest(f"For the title {title} a column can't be found")

        column_map[title] = idx + 1
    return column_map


def get_date_naive(sheet, row, col):
    value = get_value(sheet, row, col)
    if value in ("", None):
        return None
    elif not isinstance(value, Datetime):
        raise BadRequest(
            f"Problem reading {value} as a timestamp at row {row} col {col}."
        )
    return value


def get_date(sheet, row, col_name):
    dt = get_date_naive(sheet, row, col_name)
    return None if dt is None else to_utc(to_ct(dt))


def get_value(sheet, row, col):
    try:
        return sheet.cell(row=row, column=col).value
    except IndexError:
        raise BadRequest(
            f"Can't find the cell at row {row} and col {col} on sheet {sheet}."
        )


def get_str(sheet, row, col):
    value = get_value(sheet, row, col)
    if value is None:
        return None

    return value.strip()


def get_dec(sheet, row, col):
    value = get_value(sheet, row, col)
    if value in ("", None):
        return None

    try:
        return Decimal(str(value))
    except InvalidOperation as e:
        raise BadRequest(
            f"Problem parsing the number '{value}' at row {row} col {col}. {e}"
        )


def get_int(sheet, row, col):
    value = get_value(sheet, row, col)
    try:
        return int(value)
    except ValueError as e:
        raise BadRequest(
            f"Problem parsing the integer '{value}' at row {row} col {col}. {e}"
        )


ELEMENT_LOOKUP = {
    "CCL": "ccl",
    "Fixed Management Fee": "admin_fixed",
    "Gas": "admin_variable",
    "Management Fee": "admin_variable",
    "Metering and Data": "metering",
    "LDZ Customer Capacity": "dn_customer_capacity",
    "LDZ Customer Fixed": "dn_customer_fixed",
    "LDZ System Capacity": "dn_system_capacity",
    "LDZ System Commodity": "dn_system_commodity",
    "NTS Exit Capacity (ECN)": "dn_ecn",
    "NTS SO Exit": "so_exit_commodity",
    "NTS TO Exit": "to_exit_commodity",
    "Unidentified Gas": "ug",
    "WAP": "commodity",
}

QUNIT_LOOKUP = {
    "kWh": "kwh",
    "days": "days",
}


def _parse_row(bills, sheet, row, title_row):
    column_map = make_column_map(title_row)
    mprn_raw = get_value(sheet, row, column_map[Title.MPRN])
    mprn = str(mprn_raw)
    reference = str(get_value(sheet, row, column_map[Title.BILL]))
    account = get_value(sheet, row, column_map[Title.ACCOUNT])
    issue_date = get_date(sheet, row, column_map[Title.BILL_DATE])
    charge_period_from = get_date(sheet, row, column_map[Title.CHARGE_PERIOD_FROM])
    charge_period_end_naive = get_date_naive(
        sheet, row, column_map[Title.CHARGE_PERIOD_END]
    )
    if charge_period_end_naive is None:
        charge_period_end = None
    else:
        charge_period_end = to_utc(
            to_ct(charge_period_end_naive) + relativedelta(hours=23, minutes=30)
        )

    try:
        mprn_values = bills[mprn]
    except KeyError:
        mprn_values = bills[mprn] = {}

    try:
        bill = mprn_values[reference]
    except KeyError:
        bill = mprn_values[reference] = {
            "bill_type_code": "W" if reference.startswith("CR") else "N",
            "mprn": mprn,
            "reference": reference,
            "account": account,
            "issue_date": issue_date,
            "kwh": Decimal("0"),
            "breakdown": defaultdict(int, {}),
            "net_gbp": Decimal("0.00"),
            "vat_gbp": Decimal("0.00"),
            "gross_gbp": Decimal("0.00"),
            "raw_lines": [],
            "reads": [],
        }
    if charge_period_from is not None:
        if "start_date" in bill:
            bill["start_date"] = hh_min(bill["start_date"], charge_period_from)
        else:
            bill["start_date"] = charge_period_from
    if charge_period_end is not None:
        if "finish_date" in bill:
            bill["finish_date"] = hh_max(bill["finish_date"], charge_period_end)
        else:
            bill["finish_date"] = charge_period_end
    bill["raw_lines"].append([c.value for c in sheet[row]])

    bd = bill["breakdown"]
    element_desc = get_value(sheet, row, column_map[Title.CHARGE_TYPE])
    quantity = get_dec(sheet, row, column_map[Title.QUANTITY])
    qunit = get_value(sheet, row, column_map[Title.QUNIT])
    charge = get_dec(sheet, row, column_map[Title.CHARGE]) / Decimal("100")
    total = get_dec(sheet, row, column_map[Title.TOTAL])
    if bill["bill_type_code"] == "W":
        quantity = quantity * -1
    if element_desc == "VAT":
        bill["net_gbp"] += quantity
        bill["vat_gbp"] += total
        bill["gross_gbp"] += quantity + total
        if "vat_rate" not in bd:
            bd["vat_rate"] = set()
        bd["vat_rate"].add(charge)
    else:
        element_name = ELEMENT_LOOKUP[element_desc]
        if element_name == "commodity":
            bill["kwh"] += quantity
        bd[f"{element_name}_gbp"] += total
        element_qunit = QUNIT_LOOKUP[qunit]
        bd[f"{element_name}_{element_qunit}"] += quantity
        rate_name = f"{element_name}_rate"
        if rate_name not in bd:
            bd[rate_name] = set()
        bd[rate_name].add(charge)


def _make_raw_bills(sheet):
    bills = {}
    rows = tuple(sheet.rows)
    title_row = rows[0]
    for row_index, row in enumerate(rows[1:], start=2):
        val = row[0].value
        if val not in (None, ""):
            try:
                _parse_row(bills, sheet, row_index, title_row)
            except BadRequest as e:
                raise BadRequest(f"On row {row_index}: {e.description}")

    raw_bills = []
    for mprn, mprn_values in bills.items():
        for period, raw_bill in mprn_values.items():
            raw_bills.append(raw_bill)

    return raw_bills


class Parser:
    def __init__(self, f):
        self.book = load_workbook(BytesIO(f), data_only=True)
        self.sheet = self.book.worksheets[0]

        self.last_line = None
        lines = (self._set_last_line(i, l) for i, l in enumerate(f))
        self.reader = csv.reader(lines, skipinitialspace=True)
        self._line_number = None
        self._title_line = None

    @property
    def line_number(self):
        return None if self._line_number is None else self._line_number + 1

    def _set_last_line(self, i, line):
        self._line_numer = i
        self.last_line = line
        if i == 0:
            self._title_line = line
        return line

    def make_raw_bills(self):
        return _make_raw_bills(self.sheet)
