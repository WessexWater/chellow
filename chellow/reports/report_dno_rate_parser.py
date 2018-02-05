import traceback
import chellow.dloads
import os
import threading
from flask import g, request
from chellow.views import chellow_redirect
from decimal import Decimal, InvalidOperation
import openpyxl
from itertools import chain
from werkzeug.exceptions import BadRequest
from chellow.utils import req_int
from chellow.models import Session, GspGroup
from zish import dumps
from io import BytesIO
import re


def get_value(row, idx):
    val = row[idx].value
    if isinstance(val, str):
        return val.strip()
    else:
        return val


def get_rate(row, idx):
    cell = row[idx]
    if cell.number_format == '#':
        return None

    val = cell.value
    if val is None:
        return None
    elif isinstance(val, str) and len(val) == 0:
        return None
    else:
        try:
            return round(Decimal(val) / Decimal('100'), 5)
        except InvalidOperation as e:
            raise BadRequest(
                "Can't parse the decimal '" + str(val) + "' " + str(e) +
                ".") from e


def get_rag_rate(row, idx):
    val = get_rate(row, idx)
    return get_rag_rate(row, idx - 1) if val is None else val


def get_zero_rate(row, idx):
    val = get_rate(row, idx)
    return Decimal('0.00000') if val is None else val


def get_decimal(row, idx):
    return round(Decimal(get_value(row, idx)), 5)


def to_llfcs(row, idx):
    val = get_value(row, idx)
    llfcs = []
    if isinstance(val, str):
        for v in map(str.strip, val.split(',')):
            if len(v) > 0:
                if '-' in v:
                    start, finish = v.split('-')
                    for i in range(int(start), int(finish) + 1):
                        llfcs.append(str(i))
                else:
                    llfcs.append(v)
    elif isinstance(val, Decimal):
        llfcs.append(str(int(val)))
    elif isinstance(val, float):
        llfcs_str = str(int(val))
        for i in range(0, len(llfcs_str), 3):
            llfcs.append(llfcs_str[i:i+3])
    elif isinstance(val, int):
        val_str = str(val)
        for i in range(0, len(val_str), 3):
            llfcs.append(val_str[i:i+3])
    return [v.zfill(3) for v in llfcs]


VL_MAP = {
    'low voltage network': 'lv-net',
    'low-voltage network': 'lv-net',
    'low voltage substation': 'lv-sub',
    'low-voltage substation': 'lv-sub',
    'high voltage network': 'hv-net',
    'high-voltage network': 'hv-net',
    'high voltage substation': 'hv-sub',
    'high-voltage substation': 'hv-sub',
    '33kv generic': '33kv',
    '132/33kv generic': '132kv_33kv',
    '132kv generic': '132kv'}


PERIOD_MAP = {
    'peak': 'winter-weekday-peak',
    'winter': 'winter-weekday-day',
    'night': 'night',
    'other': 'other',
    'winter weekday peak': 'winter-weekday-peak',
    'winter weekday': 'winter-weekday-day'}


BAND_WEEKEND = {
    'Monday to Friday': False,
    'Weekends': True,
    'Monday to Friday (Including Bank Holidays) All Year': False,
    'Saturday and Sunday All Year': True}


def str_to_hr(hr_str):
    for sep in (':', '.'):
        if sep in hr_str:
            break
    hours, minutes = map(Decimal, hr_str.strip().split(sep))
    return hours + minutes / Decimal(60)


def val_to_slots(val):
    slots = []
    time_str = None if val is None else val.strip()
    if time_str not in (None, '', '[Start] - [End]'):
        for t_str in time_str.splitlines():
            for sep in ('-', 'to'):
                if sep in t_str:
                    break
            start_str, finish_str = t_str.split(sep)
            slots.append(
                {
                    'start': str_to_hr(start_str),
                    'finish': str_to_hr(finish_str)})
    return slots


def col_match(row, pattern, repeats=1):
    for i, cell in enumerate(row):
        txt = cell.value
        if txt is not None:
            txt_str = ' '.join(str(txt).lower().split())
            if re.search(pattern, txt_str) is not None:
                if repeats == 1:
                    return i
                else:
                    repeats -= 1

    raise BadRequest(
        "Pattern '" + pattern + "' not found in row " +
        ', '.join(str(cell.value) for cell in row))


def tab_lv_hv(sheet, gsp_rates):
    try:
        tariffs = gsp_rates['tariffs']
    except KeyError:
        tariffs = gsp_rates['tariffs'] = {}

    bands = gsp_rates['bands'] = []

    in_tariffs = False
    title_row = None
    for row in sheet.iter_rows():
        val = get_value(row, 0)
        val_0 = None if val is None else ' '.join(val.split())
        if in_tariffs:
            if val_0 is None or len(val_0) == 0:
                continue

            llfcs_str = ','.join(chain(to_llfcs(row, 1), to_llfcs(row, 10)))
            tariffs[llfcs_str] = {
                'description': val_0,
                'gbp-per-mpan-per-day': get_zero_rate(
                    row, col_match(title_row, 'fixed')),
                'gbp-per-kva-per-day': get_zero_rate(
                    row, col_match(title_row, '^capacity')),
                'excess-gbp-per-kva-per-day': get_zero_rate(
                    row, col_match(title_row, 'exce')),
                'red-gbp-per-kwh': get_rag_rate(
                    row, col_match(title_row, 'red')),
                'amber-gbp-per-kwh': get_rag_rate(
                    row, col_match(title_row, 'amber')),
                'green-gbp-per-kwh': get_rag_rate(
                    row, col_match(title_row, 'green')),
                'gbp-per-kvarh': get_zero_rate(
                    row, col_match(title_row, 'reactive'))}

        elif val_0 == 'Tariff name' or get_value(row, 1) == "Open LLFCs":
            in_tariffs = True
            title_row = row

        if val_0 in BAND_WEEKEND:
            for i, band_name in enumerate(('red', 'amber')):
                for slot in val_to_slots(get_value(row, i+1)):
                    bands.append(
                        {
                            'weekend': BAND_WEEKEND[val_0],
                            'start': slot['start'],
                            'finish': slot['finish'],
                            'band': band_name})


# State for EHV

EHV_BLANK = 0
EHV_BANDS = 1
EHV_TARIFFS = 2


def tab_ehv(sheet, gsp_rates):
    try:
        tariffs = gsp_rates['tariffs']
    except KeyError:
        tariffs = gsp_rates['tariffs'] = {}

    bands = gsp_rates['super_red'] = []

    state = EHV_BLANK
    title_row = None
    for row in sheet.iter_rows():
        val = get_value(row, 0)
        val_0 = None if val is None else ' '.join(str(val).split()).lower()
        if state == EHV_BLANK:
            if val_0 == 'time periods':
                state = EHV_BANDS
                title_row = row
            elif val_0 in (
                    'import unique identifier', 'import llfc'):
                state = EHV_TARIFFS
                title_row = row

        elif state == EHV_TARIFFS:
            for polarity, repeats in (('import', 1), ('export', 2)):
                llfc_val = get_value(
                    row, col_match(title_row, 'llfc', repeats=repeats))
                llfc = None if llfc_val is None else str(llfc_val).strip()
                if llfc not in (None, ''):
                    tariffs[llfc] = {
                        'gbp-per-kwh': get_rate(
                            row, col_match(
                                title_row, polarity + ' super red')),
                        'gbp-per-day': get_zero_rate(
                            row, col_match(title_row, polarity + ' fixed')),
                        'gbp-per-kva-per-day': get_zero_rate(
                            row, col_match(title_row, polarity + ' capacity')),
                        'excess-gbp-per-kva-per-day': get_zero_rate(
                            row, col_match(title_row, polarity + ' exce'))}

        elif state == EHV_BANDS:
            if val_0 in (None, '', 'notes'):
                state = EHV_BLANK
            else:
                period_str = ' '.join(get_value(row, 0).split()).lower()
                periods = []

                if period_str == 'monday to friday nov to feb (excluding ' + \
                        '22nd dec to 4th jan inclusive)':
                    periods.append(
                        {
                            'weekend': False,
                            'start-month': 11,
                            'start-day': 1,
                            'finish-month': 12,
                            'finish-day': 21})
                    periods.append(
                        {
                            'weekend': False,
                            'start-month': 1,
                            'start-day': 5,
                            'finish-month': 2,
                            'finish-day': 'last'})

                elif period_str in (
                        'monday to friday (including bank '
                        'holidays) november to february',
                        'monday to friday nov to feb'):
                    periods.append(
                        {
                            'weekend': False,
                            'start-month': 11,
                            'start-day': 1,
                            'finish-month': 2,
                            'finish-day': 'last'})

                for slot in val_to_slots(get_value(row, 3)):
                    for period in periods:
                        bands.append(
                            {
                                **period,
                                'start_hour': slot['start'],
                                'finish_hour': slot['finish']})


def tab_laf(sheet, gsp_rates):
    lafs = gsp_rates['lafs'] = {}
    period_lookup = {}
    for row in sheet.iter_rows():
        val_0 = get_value(row, 0)
        val_0 = None if val_0 is None else ' '.join(str(val_0).lower().split())
        if val_0 in VL_MAP:
            laf_periods = {}
            for i in range(4):
                if get_value(row, i+1) is not None:
                    laf_periods[period_lookup[i]] = get_decimal(row, i+1)

            lafs[VL_MAP[val_0]] = laf_periods
        val_1 = get_value(row, 1)
        if isinstance(val_1, str) and val_1.lower() in PERIOD_MAP:
            for i in range(4):
                key = get_value(row, i+1).lower()
                period_lookup[i] = PERIOD_MAP[key]


def content(user, file_name, file_like, gsp_group_id):
    f = sess = None
    try:
        sess = Session()
        running_name, finished_name = chellow.dloads.make_names(
            'dno_rates.zish', user)
        f = open(running_name, mode='w')
        gsp_group = GspGroup.get_by_id(sess, gsp_group_id)
        gsp_rates = {}

        if not file_name.endswith('.xlsx'):
            raise BadRequest(
                "The file extension for " + file_name + " isn't recognized.")

        book = openpyxl.load_workbook(
            file_like, data_only=True, read_only=True)

        for sheet in book.worksheets:
            title = sheet.title.strip().lower()
            if title.startswith('annex 1 '):
                tab_lv_hv(sheet, gsp_rates)
            elif title.startswith('annex 5 '):
                tab_laf(sheet, gsp_rates)
            elif title.startswith('annex 2 '):
                tab_ehv(sheet, gsp_rates)

        rs = {
            gsp_group.code: gsp_rates}

        f.write(dumps(rs))
    except BaseException:
        f.write(traceback.format_exc())
    finally:
        if sess is not None:
            sess.close()
        if f is not None:
            f.close()
            os.rename(running_name, finished_name)


def do_post(session):
    user = g.user
    file_item = request.files["dno_file"]
    gsp_group_id = req_int('gsp_group_id')

    args = user, file_item.filename, BytesIO(file_item.read()), gsp_group_id
    threading.Thread(target=content, args=args).start()
    return chellow_redirect("/downloads", 303)
