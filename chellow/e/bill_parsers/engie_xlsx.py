import csv
from datetime import datetime as Datetime
from decimal import Decimal, InvalidOperation
from enum import Enum, auto

from dateutil.relativedelta import relativedelta

from openpyxl import load_workbook
from werkzeug.exceptions import BadRequest


from chellow.utils import HH, ct_datetime, parse_mpan_core, to_ct, to_utc


ELEM_MAP = {
    None: None,
    "Charge - Recurring": {
        None: ("duos-fixed-gbp", "duos-fixed-rate", "duos-fixed-days")
    },
    "Meter - Standard": {
        "Energy Bill Relief Scheme": {None: ["ebrs-gbp", "ebrs-msp-kwh"]},
        "Energy Bill Relief Scheme Discount": {None: ["ebrs-gbp", "ebrs-msp-kwh"]},
        "Energy Bill Discount Scheme": {None: ["ebrs-gbp", "ebrs-msp-kwh"]},
    },
    "Meter - UK Electricity - AAHEDC Pass-Thru": {
        None: ["aahedc-gbp", "aahedc-rate", "aahedc-gsp-kwh"]
    },
    "Meter - UK Electricity - BSUoS Pass-Thru": {
        None: ["bsuos-gbp", "bsuos-rate", "bsuos-nbp-kwh"]
    },
    "Meter - UK Electricity - Capacity Market Pass-Thru": {
        None: ["capacity-gbp", "capacity-rate", "capacity-kwh"],
        "Reverse Capacity Market Estimate": {None: ["capacity-gbp"]},
    },
    "Meter - UK Electricity - CfD FiT Pass-Thru": {
        None: ["cfd-fit-gbp", "cfd-fit-rate", "cfd-fit-nbp-kwh"]
    },
    "Meter - UK Electricity - CCL": {
        None: ["ccl-gbp", "ccl-rate", "ccl-kwh"],
        "CCL": {None: ["ccl-gbp", "ccl-rate", "ccl-kwh"]},
        "Levy Exempt Energy": {None: ["lec-gbp", "lec-rate", "lec-kwh"]},
    },
    "Meter - UK Electricity - DUoS": {
        None: None,
        "DUoS Unit Rate 3": {
            None: ("duos-green-gbp", "duos-green-rate", "duos-green-kwh")
        },
        "DUoS Unit Charge 3": {
            None: ("duos-green-gbp", "duos-green-rate", "duos-green-kwh")
        },
        "DUoS Unit Charge 2": {
            None: ("duos-amber-gbp", "duos-amber-rate", "duos-amber-kwh")
        },
        "DUoS Unit Rate 2": {
            None: ("duos-amber-gbp", "duos-amber-rate", "duos-amber-kwh")
        },
        "DUoS Unit Charge 1": {None: ("duos-red-gbp", "duos-red-kwh", "duos-red-rate")},
        "DUoS Unit Rate 1": {None: ("duos-red-gbp", "duos-red-kwh", "duos-red-rate")},
        "DUoS Standing Charge": {
            None: ("duos-fixed-gbp", "duos-fixed-rate", "duos-fixed-days")
        },
        "DUoS Fixed": {None: ("duos-fixed-gbp", "duos-fixed-rate", "duos-fixed-days")},
        "DUoS Reactive": {
            None: ("duos-reactive-gbp", "duos-reactive-rate", "duos-reactive-kvarh")
        },
    },
    "Meter - UK Electricity - FiT Pass-Thru": {
        None: ("fit-gbp", "fit-rate", "fit-msp-kwh")
    },
    "Pass Thru - UK Electricity Cost Component": {
        None: ("meter-rental-gbp", "meter-rental-rate", "meter-rental-days")
    },
    "Meter - UK Electricity - RO Pass-Thru": {
        None: ("ro-gbp", "ro-rate", "ro-msp-kwh")
    },
    "Meter - UK Electricity - TUoS": {
        None: ("triad-gbp", "triad-rate", "triad-gsp-kw"),
        "TNUoS Fixed": {
            None: ("tnuos-gbp", "tnuos-rate", "tnuos-days"),
        },
    },
    "Meter - UK Electricity - Standard": {
        None: None,
        "Unit Rate": {
            "Summer Weekday": {
                None: (
                    "summer-weekday-gbp",
                    "summer-weekday-rate",
                    "summer-weekday-gsp-kwh",
                )
            },
            "Peak": {None: ("peak-gbp", "peak-rate", "peak-gsp-kwh")},
            "Peak Shoulder": {
                None: (
                    "peak-shoulder-gbp",
                    "peak-shoulder-gsp-kwh",
                    "peak-shoulder-rate",
                )
            },
            "Summer Night": {
                None: ("summer-night-gbp", "summer-night-rate", "summer-night-gsp-kwh")
            },
            "Summer Weekend & Bank Holiday": {
                None: (
                    "summer-weekend-gbp",
                    "summer-weekend-rate",
                    "summer-weekend-gsp-kwh",
                )
            },
            "Night": {None: ("night-gbp", "night-rate", "night-gsp-kwh")},
            "Winter Weekday": {
                None: (
                    "winter-weekday-gbp",
                    "winter-weekday-rate",
                    "winter-weekday-gsp-kwh",
                )
            },
            "Winter Weekend & Bank Holiday": {
                None: (
                    "winter-weekend-gbp",
                    "winter-weekend-rate",
                    "winter-weekend-gsp-kwh",
                )
            },
            "Winter Night": {
                None: ("winter-night-gbp", "winter-night-rate", "winter-night-gsp-kwh")
            },
            "Day": {None: ("day-gbp", "day-rate", "day-gsp-kwh")},
            "Single": {None: ("day-gbp", "day-rate", "day-gsp-kwh")},
            "Off Peak / Weekends": {None: ("day-gbp", "day-rate", "day-gsp-kwh")},
        },
        "Reverse BSUoS in Unit Rate": {
            None: ("bsuos-reverse-gbp", "bsuos-reverse-rate", "bsuos-reverse-nbp-kwh")
        },
    },
    "Meter - UK Gas - CCL": {None: ("ccl-gbp", "ccl-rate", "ccl-kwh")},
}


def _find_names(tree, path):
    if len(path) > 0:
        try:
            return _find_names(tree[path[0]], path[1:])
        except KeyError:
            pass

    return tree[None]


class Title(Enum):
    METER_POINT = auto()
    BILL_PERIOD = auto()
    FROM_DATE = auto()
    TO_DATE = auto()
    BILL_DATE = auto()
    BILL_NUMBER = auto()
    USAGE = auto()
    PRICE = auto()
    AMOUNT = auto()
    PRODUCT_ITEM_NAME = auto()
    RATE_NAME = auto()
    DESCRIPTION = auto()
    PRODUCT_ITEM_CLASS = auto()
    CUSTOMER_NUMBER = auto()


COLUMNS = {
    Title.METER_POINT: ["Meter Point", "MeterPoint"],
    Title.BILL_PERIOD: ["Bill Period", "BillPeriod"],
    Title.FROM_DATE: ["From Date", "FromDate"],
    Title.TO_DATE: ["To Date", "ToDate"],
    Title.BILL_DATE: ["Bill Date", "BillDate"],
    Title.BILL_NUMBER: ["Bill Number", "BillNumber"],
    Title.USAGE: ["Usage"],
    Title.PRICE: ["Price"],
    Title.AMOUNT: ["Amount"],
    Title.PRODUCT_ITEM_NAME: ["Product Item Name", "ProductItemName"],
    Title.RATE_NAME: ["Rate Name", "RateName"],
    Title.DESCRIPTION: ["Description"],
    Title.PRODUCT_ITEM_CLASS: ["Product Item Class", "ProductItemClass"],
    Title.CUSTOMER_NUMBER: ["Customer Number", "CustomerNumber"],
}


def make_column_map(title_row):
    titles = [cell.value for cell in title_row]
    column_map = {}
    for title, title_names in COLUMNS.items():
        idx = None
        for title_name in title_names:
            try:
                idx = titles.index(title_name)
            except ValueError:
                pass
        if idx is None:
            raise BadRequest(f"For the title {title} a column can't be found")

        column_map[title] = idx + 1
    return column_map


def get_date_naive(sheet, row, col):
    value = get_value(sheet, row, col)
    if value in ("", None):
        return None
    elif not isinstance(value, Datetime):
        raise BadRequest(
            f"Problem reading {value} as a timestamp at row {row} col {col}."
        )
    return value


def get_date(sheet, row, col_name):
    dt = get_date_naive(sheet, row, col_name)
    return None if dt is None else to_utc(to_ct(dt))


def get_value(sheet, row, col):
    try:
        return sheet.cell(row=row, column=col).value
    except IndexError:
        raise BadRequest(
            f"Can't find the cell at row {row} and col {col} on sheet {sheet}."
        )


def get_str(sheet, row, col):
    value = get_value(sheet, row, col)
    if value is None:
        return None

    return value.strip()


def get_dec(sheet, row, col):
    value = get_value(sheet, row, col)
    if value in ("", None):
        return None

    try:
        return Decimal(str(value))
    except InvalidOperation as e:
        raise BadRequest(
            f"Problem parsing the number '{value}' at row {row} col {col}. {e}"
        )


def get_int(sheet, row, col):
    value = get_value(sheet, row, col)
    try:
        return int(value)
    except ValueError as e:
        raise BadRequest(
            f"Problem parsing the integer '{value}' at row {row} col {col}. {e}"
        )


def _bd_add(bd, el_name, val):
    if el_name.split("-")[-1] in ("rate", "kva"):
        if el_name not in bd:
            bd[el_name] = set()
        bd[el_name].add(val)
    else:
        if el_name not in bd:
            bd[el_name] = 0
        try:
            bd[el_name] += val
        except TypeError as e:
            raise BadRequest(
                f"Problem with element name {el_name} and value '{val}': {e}"
            )


def _customer_mods(cust_number, description, bill):
    if cust_number == 10001596:
        if (
            description == "RO Mutualisation 2021-22"
            and "ro-gbp" in bill["breakdown"]
            and bill["issue_date"] == to_utc(ct_datetime(2023, 4, 14))
            and bill["start_date"] == to_utc(ct_datetime(2023, 3, 1))
            and bill["finish_date"] == to_utc(ct_datetime(2023, 3, 31, 23, 30))
        ):
            bill["start_date"] = to_utc(ct_datetime(2021, 4, 1))
            bill["finish_date"] = to_utc(ct_datetime(2022, 3, 31, 23, 30))

    return bill


def _parse_row(sheet, row, title_row):
    column_map = make_column_map(title_row)
    val = get_value(sheet, row, column_map[Title.METER_POINT])
    try:
        mpan_core = parse_mpan_core(str(int(val)))
    except ValueError as e:
        raise BadRequest(
            f"Can't parse the MPAN core in column 'Meter Point' with value '{val}' : "
            f"{e}"
        )

    bill_period = get_value(sheet, row, column_map[Title.BILL_PERIOD])
    if "-" in bill_period:
        period_start_naive, period_finish_naive = [
            Datetime.strptime(v, "%Y-%m-%d") for v in bill_period.split(" - ")
        ]
        period_start = to_utc(to_ct(period_start_naive))
        period_finish = to_utc(to_ct(period_finish_naive + relativedelta(days=1) - HH))
    else:
        period_start, period_finish = None, None

    from_date = get_date(sheet, row, column_map[Title.FROM_DATE])
    if from_date is None:
        if period_start is None:
            raise BadRequest("Can't find a bill start date.")
        else:
            from_date = period_start

    to_date_naive = get_date_naive(sheet, row, column_map[Title.TO_DATE])
    if to_date_naive is None:
        if period_finish is None:
            raise BadRequest("Can't find a bill finish date.")
        else:
            to_date = period_finish

    else:
        to_date = to_utc(to_ct(to_date_naive + relativedelta(days=1) - HH))

    issue_date = get_date(sheet, row, column_map[Title.BILL_DATE])
    bill_number = get_value(sheet, row, column_map[Title.BILL_NUMBER])
    bill = {
        "bill_type_code": "N",
        "kwh": Decimal(0),
        "vat": Decimal("0.00"),
        "net": Decimal("0.00"),
        "reads": [],
        "breakdown": {},
        "account": mpan_core,
        "issue_date": issue_date,
        "start_date": from_date,
        "finish_date": to_date,
        "mpan_core": mpan_core,
    }
    bd = bill["breakdown"]

    usage = get_dec(sheet, row, column_map[Title.USAGE])
    # usage_units = get_value(sheet, row, 'Usage Unit')
    price = get_dec(sheet, row, column_map[Title.PRICE])
    amount = get_dec(sheet, row, column_map[Title.AMOUNT])
    product_item_name = get_value(sheet, row, column_map[Title.PRODUCT_ITEM_NAME])
    rate_name = get_value(sheet, row, column_map[Title.RATE_NAME])
    if product_item_name == "Renewables Obligation (RO)" and usage is not None:
        bill["kwh"] += round(usage, 2)
    description = get_value(sheet, row, column_map[Title.DESCRIPTION])
    product_class = get_value(sheet, row, column_map[Title.PRODUCT_ITEM_CLASS])
    if description in ("Standard VAT@20%", "Reduced VAT@5%"):
        bill["vat"] += round(amount, 2)
        if description.endswith("20%"):
            vat_percentage = Decimal("20")
        else:
            vat_percentage = Decimal("5")
        bd["vat_percentage"] = vat_percentage
    else:
        bill["net"] += round(amount, 2)

        path = ["" if v is None else v for v in [product_class, description, rate_name]]
        names = _find_names(ELEM_MAP, path)

        duos_avail_prefix = "DUoS Availability ("
        duos_excess_avail_prefix = "DUoS Excess Availability ("

        if description.startswith("DUoS Availability Adjustment "):
            _bd_add(bd, "duos-availability-gbp", amount)
        elif description.startswith("DUoS Availability"):
            if description.startswith(duos_avail_prefix):
                _bd_add(
                    bd,
                    "duos-availability-kva",
                    int(description[len(duos_avail_prefix) : -5]),
                )
            _bd_add(bd, "duos-availability-days", usage)
            _bd_add(bd, "duos-availability-rate", price)
            _bd_add(bd, "duos-availability-gbp", amount)
        elif description.startswith("DUoS Excess Availability"):
            if description.startswith(duos_excess_avail_prefix):
                kva = int(description[len(duos_excess_avail_prefix) : -5])
                _bd_add(bd, "duos-excess-availability-kva", kva)
            _bd_add(bd, "duos-excess-availability-days", usage)
            _bd_add(bd, "duos-excess-availability-rate", price)
            _bd_add(bd, "duos-excess-availability-gbp", amount)
        elif description.startswith("BSUoS Black Start "):
            _bd_add(bd, "black-start-gbp", amount)
        elif description.startswith("BSUoS Reconciliation - "):
            if usage is not None:
                _bd_add(bd, "bsuos-nbp-kwh", usage)
            if price is not None:
                _bd_add(bd, "bsuos-rate", price)
            _bd_add(bd, "bsuos-gbp", amount)
        elif description.startswith("FiT Rec - "):
            _bd_add(bd, "fit-gbp", amount)
        elif description.startswith("FiT Reconciliation "):
            _bd_add(bd, "fit-gbp", amount)
        elif description.startswith("CfD FiT Rec - ") or description.startswith(
            "CfD FiT Reconciliation"
        ):
            _bd_add(bd, "cfd-fit-gbp", amount)
        elif description.startswith("Flex"):
            _bd_add(bd, "reconciliation-gbp", amount)
        elif description.startswith("Legacy TNUoS Reversal "):
            _bd_add(bd, "triad-gbp", amount)
        elif description.startswith("Hand Held Read -"):
            _bd_add(bd, "meter-rental-gbp", amount)
        elif description.startswith("RO Mutualisation "):
            _bd_add(bd, "ro-gbp", amount)
        elif description.startswith("OOC MOP - "):
            _bd_add(bd, "meter-rental-gbp", amount)
        elif description.startswith("KVa Adjustment "):
            _bd_add(bd, "duos-availability-gbp", amount)
        elif names is not None:
            for elem_k, elem_v in zip(names, (amount, price, usage)):
                if elem_k is not None:
                    _bd_add(bd, elem_k, elem_v)
        else:
            raise BadRequest(
                f"For the path {path} the parser can't work out the element."
            )

    reference = f"{bill_number}_{row}"
    for k, v in tuple(bd.items()):
        if isinstance(v, set):
            bd[k] = list(v)
        elif k.endswith("-gbp"):
            reference += "_" + k[:-4]

    bill["reference"] = reference
    bill["gross"] = bill["net"] + bill["vat"]
    customer_number = get_value(sheet, row, column_map[Title.CUSTOMER_NUMBER])
    return _customer_mods(customer_number, description, bill)


def _make_raw_bills(sheet):
    bills = []
    rows = tuple(sheet.rows)
    title_row = rows[0]
    for row_index, row in enumerate(rows[1:], start=2):
        if len(row) > 21:
            val = row[21].value
            if val not in (None, ""):
                try:
                    bills.append(_parse_row(sheet, row_index, title_row))
                except BadRequest as e:
                    raise BadRequest(f"On row {row_index}: {e.description}")

    return bills


class Parser:
    def __init__(self, f):
        self.book = load_workbook(f, data_only=True)
        self.sheet = self.book.worksheets[0]

        self.last_line = None
        lines = (self._set_last_line(i, l) for i, l in enumerate(f))
        self.reader = csv.reader(lines, skipinitialspace=True)
        self._line_number = None
        self._title_line = None

    @property
    def line_number(self):
        return None if self._line_number is None else self._line_number + 1

    def _set_last_line(self, i, line):
        self._line_numer = i
        self.last_line = line
        if i == 0:
            self._title_line = line
        return line

    def make_raw_bills(self):
        return _make_raw_bills(self.sheet)
