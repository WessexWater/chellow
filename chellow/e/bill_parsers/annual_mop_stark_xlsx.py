from datetime import datetime as Datetime
from decimal import Decimal, InvalidOperation

from dateutil.relativedelta import relativedelta

from openpyxl import load_workbook

from werkzeug.exceptions import BadRequest


from chellow.utils import parse_mpan_core, to_utc


def get_ct_date(row, idx):
    cell = get_cell(row, idx)
    val = cell.value
    if not isinstance(val, Datetime):
        raise BadRequest(f"Problem reading {val} as a timestamp at {cell.coordinate}.")
    return val


def get_start_date(row, idx):
    return to_utc(get_ct_date(row, idx))


def get_finish_date(row, idx):
    return to_utc(get_ct_date(row, idx) + relativedelta(hours=23, minutes=30))


def get_cell(row, idx):
    try:
        return row[idx]
    except IndexError:
        raise BadRequest(
            f"For the row {row}, the index is {idx} which is beyond the end of the row."
        )


def get_str(row, idx):
    return get_cell(row, idx).value.strip()


def get_dec(row, idx):
    cell = get_cell(row, idx)
    try:
        return Decimal(str(cell.value))
    except InvalidOperation as e:
        raise BadRequest(f"Problem parsing the number at {cell.coordinate}. {e}")


def get_int(row, idx):
    return int(get_cell(row, idx).value)


class Parser:
    def __init__(self, f):
        self.book = load_workbook(f, data_only=True)
        self.sheet = self.book.worksheets[1]

        self.last_line = None
        self._line_number = None
        self._title_line = None

    @property
    def line_number(self):
        return None if self._line_number is None else self._line_number + 1

    def _set_last_line(self, i, line):
        self._line_numer = i
        self.last_line = line
        if i == 0:
            self._title_line = line
        return line

    def make_raw_bills(self):
        try:
            bills = []
            row = next(self.sheet.iter_rows(min_row=6, max_row=6, max_col=3))
            issue_date = get_start_date(row, 2)
            for row in self.sheet.iter_rows(min_row=12, max_col=11):
                val = get_cell(row, 1).value
                if val is None or val == "":
                    break

                self._set_last_line(row[0].row, val)
                mpan_core = parse_mpan_core(str(get_int(row, 1)))
                comms = get_str(row, 2)

                settled_str = get_str(row, 3)
                if settled_str == "Settled":
                    settlement_status = "settlement"
                else:
                    settlement_status = "non_settlement"

                start_date = get_start_date(row, 5)
                finish_date = get_finish_date(row, 6)

                meter_rate = get_dec(row, 7)
                net = round(get_dec(row, 8), 2)
                vat = round(get_dec(row, 9), 2)
                gross = round(get_dec(row, 10), 2)

                breakdown = {
                    "raw-lines": [],
                    "comms": comms,
                    "settlement-status": [settlement_status],
                    "meter-rate": [meter_rate],
                    "meter-gbp": net,
                }
                bills.append(
                    {
                        "bill_type_code": "N",
                        "kwh": Decimal(0),
                        "net": net,
                        "vat": vat,
                        "gross": gross,
                        "reads": [],
                        "breakdown": breakdown,
                        "account": mpan_core,
                        "issue_date": issue_date,
                        "start_date": start_date,
                        "finish_date": finish_date,
                        "mpan_core": mpan_core,
                        "reference": "_".join(
                            (
                                start_date.strftime("%Y%m%d"),
                                finish_date.strftime("%Y%m%d"),
                                issue_date.strftime("%Y%m%d"),
                                mpan_core,
                            )
                        ),
                    }
                )
        except BadRequest as e:
            raise BadRequest(f"Row number {row[0].row}: {e.description}")

        return bills
