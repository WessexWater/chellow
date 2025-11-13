from datetime import datetime as Datetime
from decimal import Decimal, InvalidOperation

from dateutil.relativedelta import relativedelta

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from werkzeug.exceptions import BadRequest

from chellow.models import Session
from chellow.utils import parse_mpan_core, to_ct, to_utc

COPS = ("2", "3", "5", "10")

COLUMN_MAP = {
    "quarterly unmetered charge": "unmetered charge",
    "no. cop 5/10 meters": "no. cop 5 meters",
    "annual cop 5/10 dc/da rate": "annual cop 5 dc/da rate",
    "monthly cop 3 charge": "cop 3 charge",
    "cop 5/10 charge": "cop 5 charge",
    "monthly cop 5/10 charge": "cop 5 charge",
    "total": "grand total",
    "total ex vat": "grand total",
    "total in vat": "grand total 2",
    "total inc vat": "grand total 2",
    "vat @ 20 %": "vat @ 20%",
    "vat @20%": "vat @ 20%",
}
for cop in COPS:
    COLUMN_MAP[f"quarterly cop {cop} charge"] = f"cop {cop} charge"


def make_column_lookup(sheet):
    column_lookup = {}
    for cell in sheet[11]:
        title = str(cell.value).strip().lower()
        try:
            title = COLUMN_MAP[title]
        except KeyError:
            pass
        if title in column_lookup:
            title += " 2"
        column_lookup[title] = get_column_letter(cell.column)
    return column_lookup


def get_ct_date(sheet, col, row):
    cell = get_cell(sheet, col, row)
    val = cell.value
    if not isinstance(val, Datetime):
        raise BadRequest(f"Problem reading {val} as a timestamp at {cell.coordinate}.")
    return to_ct(val)


def get_start_date(sheet, col, row):
    dt = get_ct_date(sheet, col, row)
    return None if dt is None else to_utc(dt)


def get_cell(sheet, col, row):
    try:
        coordinates = f"{col}{row}"
        return sheet[coordinates]
    except IndexError:
        raise BadRequest(f"Can't find the cell {coordinates} on sheet {sheet}.")


def get_str(sheet, col, row):
    return get_cell(sheet, col, row).value.strip()


def get_dec(sheet, col, row):
    cell = get_cell(sheet, col, row)
    if cell.value is None:
        return None
    else:
        try:
            return Decimal(str(cell.value))
        except InvalidOperation as e:
            raise BadRequest(f"Problem parsing the number at {cell.coordinate}. {e}")


def get_gbp(sheet, col, row):
    gbp = get_dec(sheet, col, row)
    if gbp is None:
        return gbp
    else:
        return Decimal("0.00") + round(gbp, 2)


def get_int(sheet, col, row):
    return int(get_cell(sheet, col, row).value)


def _process_row(sess, sheet, row, issue_date, cl):
    mpan_core = parse_mpan_core(str(get_int(sheet, "B", row)))
    start_date_ct = get_ct_date(sheet, cl["bill from"], row)
    start_date = to_utc(start_date_ct)
    finish_date_ct = get_ct_date(sheet, cl["bill to"], row) + relativedelta(
        hours=23, minutes=30
    )
    finish_date = to_utc(finish_date_ct)

    elements = []
    days = (finish_date_ct - start_date_ct).days + 1

    el_titles = [("unmetered", "unmetered quant", "unmetered rate", "unmetered charge")]
    for cop in ("2", "3", "5", "10"):
        el_titles.append(
            (
                cop,
                f"no. cop {cop} meters",
                f"annual cop {cop} dc/da rate",
                f"cop {cop} charge",
            )
        )

    for cop, mpans_title, rate_title, net_title in el_titles:
        if mpans_title in cl:
            mpans = get_int(sheet, cl[mpans_title], row)
            rate = get_dec(sheet, cl[rate_title], row)
            net = get_gbp(sheet, cl[net_title], row)
            if net not in (None, 0):
                elements.append(
                    {
                        "name": "mpan",
                        "start_date": start_date,
                        "finish_date": finish_date,
                        "net": net,
                        "breakdown": {
                            "rate": {rate},
                            "days": days,
                            "mpan-days": mpans * days,
                            "cop": {cop},
                        },
                    }
                )

    for typ in ("adhoc", "regular"):
        hand_visits_title = f"no. hand held visits ({typ})"
        if hand_visits_title in cl:
            hand_visits = get_dec(sheet, cl[hand_visits_title], row)
            hand_rate = get_dec(sheet, cl[f"hand held visit ({typ}) rate"], row)
            hand_gbp = get_gbp(sheet, cl[f"hand held visit ({typ}) charge"], row)
            if hand_gbp != 0:
                elements.append(
                    {
                        "name": "activity",
                        "start_date": start_date,
                        "finish_date": finish_date,
                        "net": hand_gbp,
                        "breakdown": {
                            "rate": {hand_rate},
                            "activity-name": {"hand_held_visit"},
                            "visits": hand_visits,
                        },
                    }
                )

    annual_visits_count = get_int(sheet, cl["no. annual site visits"], row)
    annual_visits_rate = get_dec(sheet, cl["annual site visit rate"], row)
    annual_visits_gbp = get_gbp(sheet, cl["annual site visit charge"], row)
    if annual_visits_gbp != 0:
        elements.append(
            {
                "name": "activitiy",
                "start_date": start_date,
                "finish_date": finish_date,
                "net": annual_visits_gbp,
                "breakdown": {
                    "rate": {annual_visits_rate},
                    "count": annual_visits_count,
                    "activity-name": {"annual_site_visit"},
                },
            }
        )

    net = get_gbp(sheet, cl["grand total"], row)
    vat = Decimal("0.00")
    if "vat @ 20%" in cl:
        vat += get_gbp(sheet, cl["vat @ 20%"], row)
    breakdown = {
        "raw_lines": [],
        "settlement-status": ["settlement"],
        "vat": {20: {"vat": vat, "net": net}},
    }
    sum_el = sum(el["net"] for el in elements)
    if net != sum_el:
        elements.append(
            {
                "name": "activitiy",
                "start_date": start_date,
                "finish_date": finish_date,
                "net": net - sum_el,
                "breakdown": {
                    "activity-name": {"discrepancy"},
                },
            }
        )

    gross = Decimal("0.00")
    if "grand total 2" in cl:
        gross += get_gbp(sheet, cl["grand total 2"], row)

    return {
        "bill_type_code": "N",
        "kwh": Decimal(0),
        "net": net,
        "vat": vat,
        "gross": gross,
        "reads": [],
        "breakdown": breakdown,
        "account": mpan_core,
        "issue_date": issue_date,
        "start_date": start_date,
        "finish_date": finish_date,
        "mpan_core": mpan_core,
        "reference": "_".join(
            (
                start_date_ct.strftime("%Y%m%d"),
                finish_date_ct.strftime("%Y%m%d"),
                to_ct(issue_date).strftime("%Y%m%d"),
                mpan_core,
            )
        ),
        "elements": elements,
    }


class Parser:
    def __init__(self, f):
        self.book = load_workbook(f, data_only=True)
        self.sheet = self.book.worksheets[0]

        self.last_line = None
        self._line_number = None
        self._title_line = None

    @property
    def line_number(self):
        return None if self._line_number is None else self._line_number + 1

    def _set_last_line(self, i, line):
        self._line_number = i
        self.last_line = line
        if i == 0:
            self._title_line = line
        return line

    def make_raw_bills(self):
        try:
            with Session() as sess:
                bills = []
                issue_date_str = get_str(self.sheet, "A", 7)
                issue_date = to_utc(
                    to_ct(Datetime.strptime(issue_date_str[6:-3], "%d/%m/%Y %H:%M"))
                )
                column_lookup = make_column_lookup(self.sheet)

                for row in range(12, len(self.sheet["A"]) + 1):
                    val = get_cell(self.sheet, "A", row).value
                    if val is None or val == "":
                        continue

                    self._set_last_line(row, val)
                    bill = _process_row(
                        sess, self.sheet, row, issue_date, column_lookup
                    )
                    bills.append(bill)
                    sess.rollback()

        except BadRequest as e:
            raise BadRequest(f"Row number: {row} {e.description}")

        return bills
