from datetime import datetime as Datetime
from decimal import Decimal, InvalidOperation

from dateutil.relativedelta import relativedelta

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from werkzeug.exceptions import BadRequest

from chellow.models import Session
from chellow.utils import parse_mpan_core, to_ct, to_utc


COLUMN_MAP = {}


def make_column_lookup(sheet):
    column_lookup = {}
    for cell in sheet[11]:
        title = str(cell.value).strip().lower()
        try:
            title = COLUMN_MAP[title]
        except KeyError:
            pass
        if title in column_lookup:
            title += " 2"
        column_lookup[title] = get_column_letter(cell.column)
    return column_lookup


def get_ct_date(sheet, col, row):
    cell = get_cell(sheet, col, row)
    val = cell.value
    if isinstance(val, Datetime):
        pass
    elif isinstance(val, str):
        if len(val) == 10:
            val = Datetime.strptime(val, "%d/%m/%Y")
        elif len(val) == 16:
            val = Datetime.strptime(val, "%d/%m/%Y %H:%M")
        else:
            raise BadRequest(
                f"Problem parsing string {val} as a timestamp at {cell.coordinate}."
            )
    else:
        raise BadRequest(f"Problem reading {val} as a timestamp at {cell.coordinate}.")
    return to_ct(val)


def get_start_date(sheet, col, row):
    dt = get_ct_date(sheet, col, row)
    return None if dt is None else to_utc(dt)


def get_cell(sheet, col, row):
    try:
        coordinates = f"{col}{row}"
        return sheet[coordinates]
    except IndexError:
        raise BadRequest(f"Can't find the cell {coordinates} on sheet {sheet}.")


def get_str(sheet, col, row):
    return get_cell(sheet, col, row).value.strip()


def get_dec(sheet, col, row):
    cell = get_cell(sheet, col, row)
    if cell.value is None:
        return None
    else:
        try:
            return Decimal(str(cell.value))
        except InvalidOperation as e:
            raise BadRequest(f"Problem parsing the number at {cell.coordinate}. {e}")


def get_gbp(sheet, col, row):
    gbp = get_dec(sheet, col, row)
    if gbp is None:
        return gbp
    else:
        return Decimal("0.00") + round(gbp, 2)


def get_int(sheet, col, row):
    return int(get_cell(sheet, col, row).value)


def _process_row(sess, sheet, row, issue_date, cl):
    mpan_core = parse_mpan_core(str(get_int(sheet, "B", row)))
    start_date_ct = get_ct_date(sheet, cl["period from"], row)
    start_date = to_utc(start_date_ct)
    finish_date_ct = get_ct_date(sheet, cl["period to"], row) + relativedelta(
        hours=23, minutes=30
    )
    finish_date = to_utc(finish_date_ct)

    elements = []
    days = (finish_date_ct - start_date_ct).days + 1

    rate = get_dec(sheet, cl["dc pa"], row)
    net = get_gbp(sheet, cl["charge"], row)
    settlement_status_str = get_str(sheet, cl["type"], row)
    settlement_status = (
        "settlement" if settlement_status_str == "Settled" else "non-settlement"
    )
    comm = get_str(sheet, cl["comms"], row)
    elements.append(
        {
            "name": "mpan",
            "start_date": start_date,
            "finish_date": finish_date,
            "net": net,
            "breakdown": {
                "settlement-status": {settlement_status},
                "rate": {rate},
                "days": days,
                "comm": {comm},
            },
        }
    )

    vat = get_gbp(sheet, cl["est. vat"], row)
    gross = get_gbp(sheet, cl["est. total"], row)

    breakdown = {
        "raw_lines": [],
        "vat": {20: {"vat": vat, "net": net}},
    }

    return {
        "bill_type_code": "N",
        "kwh": Decimal(0),
        "net": net,
        "vat": vat,
        "gross": gross,
        "reads": [],
        "breakdown": breakdown,
        "account": mpan_core,
        "issue_date": issue_date,
        "start_date": start_date,
        "finish_date": finish_date,
        "mpan_core": mpan_core,
        "reference": "_".join(
            (
                start_date_ct.strftime("%Y%m%d"),
                finish_date_ct.strftime("%Y%m%d"),
                to_ct(issue_date).strftime("%Y%m%d"),
                mpan_core,
            )
        ),
        "elements": elements,
    }


class Parser:
    def __init__(self, f):
        self.book = load_workbook(f, data_only=True)
        self.sheet = self.book.worksheets[1]

        self.last_line = None
        self._line_number = None
        self._title_line = None

    @property
    def line_number(self):
        return None if self._line_number is None else self._line_number + 1

    def _set_last_line(self, i, line):
        self._line_number = i
        self.last_line = line
        if i == 0:
            self._title_line = line
        return line

    def make_raw_bills(self):
        try:
            with Session() as sess:
                column_lookup = make_column_lookup(self.sheet)
                bills = []
                issue_date = get_start_date(self.sheet, "C", 6)

                for row in range(12, len(self.sheet["B"]) + 1):
                    val = get_cell(self.sheet, "B", row).value
                    if val is None or val == "":
                        continue

                    self._set_last_line(row, val)
                    bill = _process_row(
                        sess, self.sheet, row, issue_date, column_lookup
                    )
                    bills.append(bill)
                    sess.rollback()

        except BadRequest as e:
            raise BadRequest(f"Row number: {row} {e.description}")

        return bills
