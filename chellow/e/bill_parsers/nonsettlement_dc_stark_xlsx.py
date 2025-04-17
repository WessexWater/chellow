from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from werkzeug.exceptions import BadRequest

from chellow.e.computer import hh_rate
from chellow.models import Session
from chellow.utils import ct_datetime, parse_mpan_core, to_utc


def get_ct_date(title_row, row, name):
    return get_value(title_row, row, name)


def get_start_date(title_row, row, name):
    return to_utc(get_ct_date(title_row, row, name))


def get_value(title_row, row, name):
    idx = None
    name = name.strip().lower()

    for i, title_cell in enumerate(title_row):
        if str(title_cell.value).strip().lower() == name:
            idx = i
            break

    if idx is None:
        raise BadRequest(
            "The name '{name}' can't be found in the titles "
            "{title_row}.".format(name=name, title_row=title_row)
        )

    try:
        return row[idx].value

    except IndexError:
        raise BadRequest(
            f"For the row {row}, the index is {idx} which is beyond the end of the row."
        )


def get_str(title_row, row, name):
    return get_value(title_row, row, name).strip()


def get_dec(title_row, row, name):
    try:
        return Decimal(str(get_value(title_row, row, name)))
    except InvalidOperation:
        return None


def get_int(title_row, row, name):
    return int(get_value(title_row, row, name))


class Parser:
    def __init__(self, f):
        self.book = load_workbook(f)
        self.sheet = self.book.worksheets[0]

        self.last_line = None
        self._line_number = None
        self._title_line = None

    @property
    def line_number(self):
        return None if self._line_number is None else self._line_number + 1

    def _set_last_line(self, i, line):
        self._line_number = i
        self.last_line = line
        if i == 0:
            self._title_line = line
        return line

    def make_raw_bills(self):
        row_index = None
        caches = {}
        try:
            with Session() as sess:
                bills = []
                row_iter = self.sheet.iter_rows()
                title_row = next(row_iter)
                for row in row_iter:
                    val = get_value(title_row, row, "mpan ref")
                    if val is None or val == "":
                        break

                    self._set_last_line(row_index, val)
                    msn = str(get_value(title_row, row, "meter")).strip()
                    mpan_core = parse_mpan_core(
                        str(get_int(title_row, row, "mpan ref"))
                    )
                    start_date_ct = get_ct_date(title_row, row, "start")
                    start_date = to_utc(start_date_ct)
                    issue_date = start_date
                    finish_date_ct = get_ct_date(title_row, row, "end")
                    finish_date = to_utc(
                        ct_datetime(
                            finish_date_ct.year,
                            finish_date_ct.month,
                            finish_date_ct.day,
                            23,
                            30,
                        )
                    )
                    check = get_str(title_row, row, "check")
                    if check != "Billed":
                        continue
                    rates = hh_rate(sess, caches, 0, start_date)
                    meter_rate = rates["annual_rates"]["non_settlement"]["*"]["IP"][
                        "*"
                    ]["gbp_per_meter"]
                    months = (finish_date_ct.year - start_date_ct.year) * 12 + (
                        finish_date_ct.month - start_date_ct.month + 1
                    )
                    net = round(Decimal(float(meter_rate) / 12 * months), 2)
                    vat = round(net * Decimal("0.2"), 2)

                    breakdown = {
                        "raw_lines": [],
                        "cop": ["5"],
                        "settlement-status": ["non_settlement"],
                        "msn": [msn],
                        "meter-rate": [meter_rate],
                        "months": months,
                        "meter-gbp": net,
                    }

                    bills.append(
                        {
                            "bill_type_code": "N",
                            "kwh": Decimal(0),
                            "vat": vat,
                            "net": net,
                            "gross": net + vat,
                            "reads": [],
                            "breakdown": breakdown,
                            "account": mpan_core,
                            "issue_date": issue_date,
                            "start_date": start_date,
                            "finish_date": finish_date,
                            "mpan_core": mpan_core,
                            "reference": "_".join(
                                (
                                    start_date.strftime("%Y%m%d"),
                                    finish_date.strftime("%Y%m%d"),
                                    issue_date.strftime("%Y%m%d"),
                                    mpan_core,
                                )
                            ),
                        }
                    )
                    sess.rollback()
        except BadRequest as e:
            raise BadRequest(f"Row number: {row_index} {e.description}")

        return bills
