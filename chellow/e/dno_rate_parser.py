import re
from decimal import Decimal, InvalidOperation
from enum import Enum, auto
from io import BytesIO
from itertools import chain
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from sqlalchemy import null, or_, select

from werkzeug.exceptions import BadRequest

from chellow.models import Contract, Llfc, Party, RateScript, VoltageLevel
from chellow.rate_server import download
from chellow.utils import ct_datetime, hh_format, to_utc


def get_cell(sheet, col, row):
    try:
        coordinates = f"{col}{row}"
        return sheet[coordinates]
    except IndexError:
        raise BadRequest(f"Can't find the cell {coordinates} on sheet {sheet}.")


def get_int(sheet, col, row):
    return int(get_cell(sheet, col, row).value)


def get_dec(sheet, col, row):
    cell = get_cell(sheet, col, row)
    try:
        return Decimal(str(cell.value))
    except InvalidOperation as e:
        raise BadRequest(f"Problem parsing the number at {cell.coordinate}. {e}")


def get_str(sheet, col, row):
    return get_cell(sheet, col, row).value.strip()


def get_value(row, idx):
    val = row[idx].value
    if isinstance(val, str):
        return val.strip()
    else:
        return val


def get_rate(row, idx):
    cell = row[idx]
    if cell.number_format == "#":
        return None

    val = cell.value
    if val is None:
        return None
    elif isinstance(val, str) and len(val) == 0:
        return None
    else:
        try:
            return round(Decimal(val) / Decimal("100"), 5)
        except InvalidOperation as e:
            raise BadRequest(f"Can't parse the decimal '{val}' {e}.") from e


def get_rag_rate(row, idx):
    val = get_rate(row, idx)
    return get_rag_rate(row, idx - 1) if val is None else val


def get_zero_rate(row, idx):
    val = get_rate(row, idx)
    return Decimal("0.00000") if val is None else val


def get_decimal(row, idx):
    return round(Decimal(get_value(row, idx)), 5)


def to_llfcs(val):
    llfcs = []
    if isinstance(val, str):
        val = ",".join(val.splitlines())
        val = val.replace("/", ",").replace("inclusive", ",")
        for v in map(str.strip, val.split(",")):
            if len(v) > 0:
                if "-" in v:
                    start, finish = v.split("-")
                    m = re.search(r"\d", start)
                    s = m.start()
                    m = re.search(r"\d", finish)
                    f = m.start()

                    for i in range(int(start[s:]), int(finish[f:]) + 1):
                        suff = str(i).zfill(3 - s)
                        llfcs.append(f"{start[:s]}{suff}")
                else:
                    llfcs.append(v)

    elif isinstance(val, Decimal):
        llfcs.append(str(int(val)))
    elif isinstance(val, float):
        llfcs_str = str(int(val))
        for i in range(0, len(llfcs_str), 3):
            llfcs.append(llfcs_str[i : i + 3])
    elif isinstance(val, int):
        val_str = str(val)
        for i in range(0, len(val_str), 3):
            llfcs.append(val_str[i : i + 3])
    return [v.zfill(3) for v in llfcs]


def to_pcs(row, idx):
    val = get_value(row, idx)
    llfcs = []
    if isinstance(val, str):
        for v in map(str.strip, val.replace("or", ",").replace("to", "-").split(",")):
            if len(v) > 0:
                if "-" in v:
                    start, finish = v.split("-")
                    for i in range(int(start), int(finish) + 1):
                        llfcs.append(str(i))
                else:
                    llfcs.append(v)
    elif isinstance(val, Decimal):
        llfcs.append(str(int(val)))
    elif isinstance(val, float):
        llfcs_str = str(int(val))
        for i in range(0, len(llfcs_str), 3):
            llfcs.append(llfcs_str[i : i + 3])
    elif isinstance(val, int):
        val_str = str(val)
        for i in range(0, len(val_str), 3):
            llfcs.append(val_str[i : i + 3])
    return [v.zfill(2) for v in llfcs]


BAND_WEEKEND = {
    "monday to friday": False,
    "weekends": True,
    "monday to friday (including bank holidays) all year": False,
    "saturday and sunday all year": True,
    "notes": None,
}


def str_to_hr(hr_str):
    for sep in (":", "."):
        if sep in hr_str:
            break

    time_strs_raw = hr_str.strip().split(sep)
    len_time_strs_raw = len(time_strs_raw)
    if len_time_strs_raw == 1:
        ts = time_strs_raw[0]
        time_strs = ts[:2], ts[2:]
    elif len_time_strs_raw == 2:
        time_strs = time_strs_raw
    else:
        raise BadRequest(f"Can't work out the hours and minutes from '{hr_str}'")

    hours, minutes = map(Decimal, time_strs)
    return hours + minutes / Decimal(60)


def val_to_slots(val):
    slots = []
    time_str = None if val is None else str(val).strip()
    if time_str not in (None, "", "[Start] - [End]", "0"):
        for t_str in time_str.splitlines():
            for sep in ("-", "to"):
                if sep in t_str:
                    break
            start_str, finish_str = t_str.split(sep)
            slots.append(
                {"start": str_to_hr(start_str), "finish": str_to_hr(finish_str)}
            )
    return slots


def col_match(row, pattern, repeats=1):
    for i, cell in enumerate(row):
        txt = cell.value
        if txt is not None:
            txt_str = " ".join(str(txt).lower().split())
            if re.search(pattern, txt_str) is not None:
                if repeats == 1:
                    return i
                else:
                    repeats -= 1

    raise BadRequest(
        f"Pattern '{pattern}' not found in row "
        + ", ".join(str(cell.value) for cell in row)
    )


def tab_lv_hv(sheet, gsp_rates):
    try:
        tariffs = gsp_rates["tariffs"]
    except KeyError:
        tariffs = gsp_rates["tariffs"] = {}

    bands = gsp_rates["bands"] = []

    class State(Enum):
        OUTSIDE = auto()
        BANDS = auto()
        TARIFFS = auto()

    state = State.OUTSIDE
    title_row = None
    for row in sheet.iter_rows():
        val = get_value(row, 0)
        val_0 = None if val is None else " ".join(val.split())
        val_0_lower = None if val is None else val_0.lower()
        if state == State.OUTSIDE:
            if val_0_lower == "tariff name" or get_value(row, 1) == "Open LLFCs":
                state = State.TARIFFS
                title_row = row
            elif val_0_lower == "time periods":
                state = State.BANDS

        elif state == State.BANDS:
            try:
                is_weekend = BAND_WEEKEND[val_0_lower]
            except KeyError:
                raise BadRequest(f"The key '{val_0_lower}' isn't recognised.")

            if is_weekend is None:
                state = State.OUTSIDE
            else:
                for i, band_name in enumerate(("red", "amber")):
                    for slot in val_to_slots(get_value(row, i + 1)):
                        bands.append(
                            {
                                "weekend": is_weekend,
                                "start": slot["start"],
                                "finish": slot["finish"],
                                "band": band_name,
                            }
                        )

        elif state == State.TARIFFS:
            if val_0 is None or len(val_0) == 0:
                state = State.OUTSIDE
            else:
                llfcs_str = ",".join(
                    chain(to_llfcs(get_value(row, 1)), to_llfcs(get_value(row, 10)))
                )
                pcs_str = ",".join(to_pcs(row, 2))
                tariffs[llfcs_str + "_" + pcs_str] = {
                    "description": val_0,
                    "gbp-per-mpan-per-day": get_zero_rate(
                        row, col_match(title_row, "fixed")
                    ),
                    "gbp-per-kva-per-day": get_zero_rate(
                        row, col_match(title_row, "^capacity")
                    ),
                    "excess-gbp-per-kva-per-day": get_zero_rate(
                        row, col_match(title_row, "exce")
                    ),
                    "red-gbp-per-kwh": get_rag_rate(row, col_match(title_row, "red")),
                    "amber-gbp-per-kwh": get_rag_rate(
                        row, col_match(title_row, "amber")
                    ),
                    "green-gbp-per-kwh": get_rag_rate(
                        row, col_match(title_row, "green")
                    ),
                    "gbp-per-kvarh": get_zero_rate(
                        row, col_match(title_row, "reactive")
                    ),
                }


# State for EHV

EHV_BLANK = 0
EHV_BANDS = 1
EHV_TARIFFS = 2


def tab_ehv(sheet, gsp_rates):
    try:
        tariffs = gsp_rates["tariffs"]
    except KeyError:
        tariffs = gsp_rates["tariffs"] = {}

    bands = gsp_rates["super_red"] = []

    state = EHV_BLANK
    title_row = None
    for row in sheet.iter_rows():
        val = get_value(row, 0)
        val_0 = None if val is None else " ".join(str(val).split()).lower()
        if state == EHV_BLANK:
            if val_0 == "time periods":
                state = EHV_BANDS
                title_row = row
            elif val_0 in ("import unique identifier", "import llfc"):
                state = EHV_TARIFFS
                title_row = row

        elif state == EHV_TARIFFS:
            for polarity, repeats in (("import", 1), ("export", 2)):
                llfc_val = get_value(row, col_match(title_row, "llfc", repeats=repeats))
                llfc = None if llfc_val is None else str(llfc_val).strip()
                if llfc not in (None, ""):
                    tariffs[llfc] = {
                        "gbp-per-kwh": get_rate(
                            row, col_match(title_row, polarity + " super red")
                        ),
                        "gbp-per-day": get_zero_rate(
                            row, col_match(title_row, polarity + " fixed")
                        ),
                        "gbp-per-kva-per-day": get_zero_rate(
                            row, col_match(title_row, polarity + " capacity")
                        ),
                        "excess-gbp-per-kva-per-day": get_zero_rate(
                            row, col_match(title_row, polarity + " exce")
                        ),
                    }

        elif state == EHV_BANDS:
            if val_0 in (None, "", "notes"):
                state = EHV_BLANK
            else:
                period_str = " ".join(get_value(row, 0).split()).lower()
                periods = []

                if (
                    period_str
                    == "monday to friday nov to feb (excluding "
                    + "22nd dec to 4th jan inclusive)"
                ):
                    periods.append(
                        {
                            "weekend": False,
                            "start-month": 11,
                            "start-day": 1,
                            "finish-month": 12,
                            "finish-day": 21,
                        }
                    )
                    periods.append(
                        {
                            "weekend": False,
                            "start-month": 1,
                            "start-day": 5,
                            "finish-month": 2,
                            "finish-day": "last",
                        }
                    )

                elif period_str in (
                    "monday to friday (including bank holidays) november to february",
                    "monday to friday nov to feb",
                ):
                    periods.append(
                        {
                            "weekend": False,
                            "start-month": 11,
                            "start-day": 1,
                            "finish-month": 2,
                            "finish-day": "last",
                        }
                    )

                for slot in val_to_slots(get_value(row, 3)):
                    for period in periods:
                        bands.append(
                            {
                                **period,
                                "start_hour": slot["start"],
                                "finish_hour": slot["finish"],
                            }
                        )


def find_rates(file_name, file_like):
    rates = {"a_file_name": file_name}
    vls = []

    if file_name.endswith(".zip"):
        try:
            with ZipFile(file_like) as za:
                for zname in za.namelist():
                    if zname.endswith(".xlsx"):
                        with za.open(zname) as f:
                            gsp_code, gsp_rates, gsp_vls = find_gsp_group_rates(
                                zname, BytesIO(f.read())
                            )
                            rates[gsp_code] = gsp_rates
                            vls.extend(gsp_vls)
        except BadZipFile as e:
            raise BadRequest(f"Problem with zip file '{file_name}'") from e

    elif file_name.endswith(".xlsx"):
        gsp_code, gsp_rates, gsp_vls = find_gsp_group_rates(file_name, file_like)
        rates[gsp_code] = gsp_rates
        vls.extend(gsp_vls)

    else:
        raise BadRequest(f"The file extension for {file_name} isn't recognized.")

    return rates, vls


GSP_MAP = (
    ("gsp a", "_A"),
    ("gsp b", "_B"),
    ("gsp c", "_C"),
    ("gsp d", "_D"),
    ("gsp e", "_E"),
    ("gsp f", "_F"),
    ("gsp g", "_G"),
    ("gsp h", "_H"),
    ("gsp j", "_J"),
    ("gsp k", "_K"),
    ("gsp l", "_L"),
    ("gsp m", "_M"),
    ("gsp n", "_N"),
    ("gsp p", "_P"),
    ("gsp_a", "_A"),
    ("gsp_b", "_B"),
    ("gsp_c", "_C"),
    ("gsp_d", "_D"),
    ("gsp_e", "_E"),
    ("gsp_f", "_F"),
    ("gsp_g", "_G"),
    ("gsp_h", "_H"),
    ("gsp_j", "_J"),
    ("gsp_k", "_K"),
    ("gsp_l", "_L"),
    ("gsp_m", "_M"),
    ("gsp_n", "_N"),
    ("gsp_p", "_P"),
    ("mide", "_E"),
    ("sepd", "_H"),
    ("sweb", "_L"),
)


def find_gsp_group_rates(file_name, file_like):
    rates = {}

    fname = file_name.lower()
    gsp_code = None
    for substr, code in GSP_MAP:
        if substr in fname:
            gsp_code = code
            break

    if gsp_code is None:
        raise BadRequest(
            f"Can't determine the GSP group from the file name '{file_name}'."
        )

    book = load_workbook(file_like, data_only=True)

    vls = None

    try:
        for sheet in book.worksheets:
            title = sheet.title.strip().lower()
            if title.startswith("annex 1 "):
                tab_lv_hv(sheet, rates)
            elif title.startswith("annex 2 "):
                tab_ehv(sheet, rates)
            elif title.startswith("annex 5 "):
                vls = tab_llfs(sheet)
    except BadRequest as e:
        raise BadRequest(f"Problem with file '{file_name}': {e.description}")

    return gsp_code, rates, vls


def rate_server_import(sess, s, paths, logger):
    logger("Starting to check for new DNO spreadsheets")
    year_entries = {}
    for path, url in paths:
        if len(path) == 5:

            year_str, utility, rate_type, dno_code, file_name = path
            year = int(year_str)
            try:
                dno_entries = year_entries[year]
            except KeyError:
                dno_entries = year_entries[year] = {}

            if utility == "electricity" and rate_type == "duos":

                try:
                    fl_entries = dno_entries[dno_code]
                except KeyError:
                    fl_entries = dno_entries[dno_code] = {}

                fl_entries[file_name] = url

    for year, dno_entries in sorted(year_entries.items()):
        fy_start = to_utc(ct_datetime(year, 4, 1))
        for dno_code, fl_entries in sorted(dno_entries.items()):
            contract = Contract.get_dno_by_name(sess, dno_code)
            if fy_start < contract.start_rate_script.start_date:
                continue

            rs = sess.execute(
                select(RateScript).where(
                    RateScript.contract == contract,
                    RateScript.start_date == fy_start,
                )
            ).scalar_one_or_none()
            if rs is None:
                rs = contract.insert_rate_script(sess, fy_start, {})

            file_name, url = sorted(fl_entries.items())[-1]

            rs_script = rs.make_script()
            if rs_script.get("a_file_name") != file_name:
                try:
                    fl = BytesIO(download(s, url))
                    rates, vls = find_rates(file_name, fl)
                    rs.update(rates)
                    logger(
                        f"Updated DNO {dno_code} rate script for "
                        f"{hh_format(fy_start)}"
                    )
                    update_vls(sess, logger, vls, dno_code, fy_start, rs.finish_date)
                except BadRequest as e:
                    raise BadRequest(
                        f"Problem with year {year} DNO {dno_code}: {e.description}"
                    )

    logger("Finished DNO spreadsheets")
    sess.commit()


LV_NET = ("LV", False)
LV_SUB = ("LV", True)
HV_NET = ("HV", False)
HV_SUB = ("HV", True)
EHV_NET = ("EHV", False)


VL_LOOKUP = {
    "132/33kV generic": EHV_NET,
    "132/EHV connected": EHV_NET,
    "132/HV connected": HV_NET,
    "132kV connected": EHV_NET,
    "132kV generic": EHV_NET,
    "132kV generic (demand)": EHV_NET,
    "132kV generic (generation)": EHV_NET,
    "132kV generic Export": EHV_NET,
    "132kV generic Import": EHV_NET,
    "132kV to 33kV generic": EHV_NET,
    "33kV generic": EHV_NET,
    "33kV generic (demand)": EHV_NET,
    "33kV generic (generation)": EHV_NET,
    "33kV generic Export": EHV_NET,
    "33kV generic Import": EHV_NET,
    "EHV 33kV Export": EHV_NET,
    "EHV 33kV Import": EHV_NET,
    "EHV connected": EHV_NET,
    "Greater than 22kV connected - demand": HV_NET,
    "Greater than 22kV connected - generation": HV_NET,
    "High-voltage network": HV_NET,
    "High Voltage Network": HV_NET,
    "High-voltage substation": HV_SUB,
    "High Voltage Substation": HV_SUB,
    "Low-voltage network": LV_NET,
    "Low Voltage Network": LV_NET,
    "Low-voltage substation": LV_SUB,
    "Low Voltage Substation": LV_SUB,
}


def tab_llfs(sheet):
    llfcs = []

    in_llfcs = False
    for row in range(1, len(sheet["A"]) + 1):
        val = get_cell(sheet, "A", row).value
        val_0 = None if val is None else " ".join(val.split())
        if in_llfcs:
            if val_0 is None or len(val_0) == 0:
                in_llfcs = False
            else:
                vl, is_substation = VL_LOOKUP[val_0]
                for llfc_code in to_llfcs(get_cell(sheet, "H", row).value):
                    llfc = {
                        "code": llfc_code,
                        "voltage_level": vl,
                        "is_substation": is_substation,
                    }
                    llfcs.append(llfc)

        elif val_0 == "Metered voltage":
            in_llfcs = True

    return llfcs


def update_vls(sess, logger, vls, dno_code, fy_start, rs_finish):
    dno = Party.get_dno_by_code(sess, dno_code, fy_start)

    for vl in vls:
        vl_code = vl["code"]
        q = select(Llfc).where(
            Llfc.dno == dno,
            Llfc.code == vl_code,
            or_(Llfc.valid_to == null(), Llfc.valid_to >= fy_start),
        )
        if rs_finish is not None:
            q = q.where(Llfc.valid_from <= rs_finish)
        llfc = sess.execute(q).scalar_one_or_none()

        if llfc is None:
            raise BadRequest(
                f"There is no LLFC with the code '{vl_code}' associated with the DNO "
                f"{dno.code} from {hh_format(fy_start)} to {hh_format(rs_finish)}."
            )

        vl_voltage_level = VoltageLevel.get_by_code(sess, vl["voltage_level"])
        llfc.voltage_level = vl_voltage_level

        llfc.is_substation = vl["is_substation"]
        if sess.is_modified(llfc):
            logger(f"Updated LLFC {llfc.code} of DNO {dno_code}")
            sess.flush()
