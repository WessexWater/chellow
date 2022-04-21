import re
from decimal import Decimal, InvalidOperation
from itertools import chain

import openpyxl

from werkzeug.exceptions import BadRequest


def get_value(row, idx):
    val = row[idx].value
    if isinstance(val, str):
        return val.strip()
    else:
        return val


def get_rate(row, idx):
    cell = row[idx]
    if cell.number_format == "#":
        return None

    val = cell.value
    if val is None:
        return None
    elif isinstance(val, str) and len(val) == 0:
        return None
    else:
        try:
            return round(Decimal(val) / Decimal("100"), 5)
        except InvalidOperation as e:
            raise BadRequest(f"Can't parse the decimal '{val}' {e}.") from e


def get_rag_rate(row, idx):
    val = get_rate(row, idx)
    return get_rag_rate(row, idx - 1) if val is None else val


def get_zero_rate(row, idx):
    val = get_rate(row, idx)
    return Decimal("0.00000") if val is None else val


def get_decimal(row, idx):
    return round(Decimal(get_value(row, idx)), 5)


def to_llfcs(row, idx):
    val = get_value(row, idx)
    llfcs = []
    if isinstance(val, str):
        for v in map(str.strip, val.split(",")):
            if len(v) > 0:
                if "-" in v:
                    start, finish = v.split("-")
                    m = re.search(r"\d", start)
                    s = m.start()
                    m = re.search(r"\d", finish)
                    f = m.start()

                    for i in range(int(start[s:]), int(finish[f:]) + 1):
                        suff = str(i).zfill(3 - s)
                        llfcs.append(f"{start[:s]}{suff}")
                else:
                    llfcs.append(v)

    elif isinstance(val, Decimal):
        llfcs.append(str(int(val)))
    elif isinstance(val, float):
        llfcs_str = str(int(val))
        for i in range(0, len(llfcs_str), 3):
            llfcs.append(llfcs_str[i : i + 3])
    elif isinstance(val, int):
        val_str = str(val)
        for i in range(0, len(val_str), 3):
            llfcs.append(val_str[i : i + 3])
    return [v.zfill(3) for v in llfcs]


def to_pcs(row, idx):
    val = get_value(row, idx)
    llfcs = []
    if isinstance(val, str):
        for v in map(str.strip, val.replace("or", ",").replace("to", "-").split(",")):
            if len(v) > 0:
                if "-" in v:
                    start, finish = v.split("-")
                    for i in range(int(start), int(finish) + 1):
                        llfcs.append(str(i))
                else:
                    llfcs.append(v)
    elif isinstance(val, Decimal):
        llfcs.append(str(int(val)))
    elif isinstance(val, float):
        llfcs_str = str(int(val))
        for i in range(0, len(llfcs_str), 3):
            llfcs.append(llfcs_str[i : i + 3])
    elif isinstance(val, int):
        val_str = str(val)
        for i in range(0, len(val_str), 3):
            llfcs.append(val_str[i : i + 3])
    return [v.zfill(2) for v in llfcs]


BAND_WEEKEND = {
    "Monday to Friday": False,
    "Weekends": True,
    "Monday to Friday (Including Bank Holidays) All Year": False,
    "Saturday and Sunday All Year": True,
}


def str_to_hr(hr_str):
    for sep in (":", "."):
        if sep in hr_str:
            break
    hours, minutes = map(Decimal, hr_str.strip().split(sep))
    return hours + minutes / Decimal(60)


def val_to_slots(val):
    slots = []
    time_str = None if val is None else str(val).strip()
    if time_str not in (None, "", "[Start] - [End]", "0"):
        for t_str in time_str.splitlines():
            for sep in ("-", "to"):
                if sep in t_str:
                    break
            start_str, finish_str = t_str.split(sep)
            slots.append(
                {"start": str_to_hr(start_str), "finish": str_to_hr(finish_str)}
            )
    return slots


def col_match(row, pattern, repeats=1):
    for i, cell in enumerate(row):
        txt = cell.value
        if txt is not None:
            txt_str = " ".join(str(txt).lower().split())
            if re.search(pattern, txt_str) is not None:
                if repeats == 1:
                    return i
                else:
                    repeats -= 1

    raise BadRequest(
        f"Pattern '{pattern}' not found in row "
        + ", ".join(str(cell.value) for cell in row)
    )


def tab_lv_hv(sheet, gsp_rates):
    try:
        tariffs = gsp_rates["tariffs"]
    except KeyError:
        tariffs = gsp_rates["tariffs"] = {}

    bands = gsp_rates["bands"] = []

    in_tariffs = False
    title_row = None
    for row in sheet.iter_rows():
        val = get_value(row, 0)
        val_0 = None if val is None else " ".join(val.split())
        if in_tariffs:
            if val_0 is None or len(val_0) == 0:
                in_tariffs = False
            else:
                llfcs_str = ",".join(chain(to_llfcs(row, 1), to_llfcs(row, 10)))
                pcs_str = ",".join(to_pcs(row, 2))
                tariffs[llfcs_str + "_" + pcs_str] = {
                    "description": val_0,
                    "gbp-per-mpan-per-day": get_zero_rate(
                        row, col_match(title_row, "fixed")
                    ),
                    "gbp-per-kva-per-day": get_zero_rate(
                        row, col_match(title_row, "^capacity")
                    ),
                    "excess-gbp-per-kva-per-day": get_zero_rate(
                        row, col_match(title_row, "exce")
                    ),
                    "red-gbp-per-kwh": get_rag_rate(row, col_match(title_row, "red")),
                    "amber-gbp-per-kwh": get_rag_rate(
                        row, col_match(title_row, "amber")
                    ),
                    "green-gbp-per-kwh": get_rag_rate(
                        row, col_match(title_row, "green")
                    ),
                    "gbp-per-kvarh": get_zero_rate(
                        row, col_match(title_row, "reactive")
                    ),
                }

        elif val_0 == "Tariff name" or get_value(row, 1) == "Open LLFCs":
            in_tariffs = True
            title_row = row

        if val_0 in BAND_WEEKEND:
            for i, band_name in enumerate(("red", "amber")):
                for slot in val_to_slots(get_value(row, i + 1)):
                    bands.append(
                        {
                            "weekend": BAND_WEEKEND[val_0],
                            "start": slot["start"],
                            "finish": slot["finish"],
                            "band": band_name,
                        }
                    )


# State for EHV

EHV_BLANK = 0
EHV_BANDS = 1
EHV_TARIFFS = 2


def tab_ehv(sheet, gsp_rates):
    try:
        tariffs = gsp_rates["tariffs"]
    except KeyError:
        tariffs = gsp_rates["tariffs"] = {}

    bands = gsp_rates["super_red"] = []

    state = EHV_BLANK
    title_row = None
    for row in sheet.iter_rows():
        val = get_value(row, 0)
        val_0 = None if val is None else " ".join(str(val).split()).lower()
        if state == EHV_BLANK:
            if val_0 == "time periods":
                state = EHV_BANDS
                title_row = row
            elif val_0 in ("import unique identifier", "import llfc"):
                state = EHV_TARIFFS
                title_row = row

        elif state == EHV_TARIFFS:
            for polarity, repeats in (("import", 1), ("export", 2)):
                llfc_val = get_value(row, col_match(title_row, "llfc", repeats=repeats))
                llfc = None if llfc_val is None else str(llfc_val).strip()
                if llfc not in (None, ""):
                    tariffs[llfc] = {
                        "gbp-per-kwh": get_rate(
                            row, col_match(title_row, polarity + " super red")
                        ),
                        "gbp-per-day": get_zero_rate(
                            row, col_match(title_row, polarity + " fixed")
                        ),
                        "gbp-per-kva-per-day": get_zero_rate(
                            row, col_match(title_row, polarity + " capacity")
                        ),
                        "excess-gbp-per-kva-per-day": get_zero_rate(
                            row, col_match(title_row, polarity + " exce")
                        ),
                    }

        elif state == EHV_BANDS:
            if val_0 in (None, "", "notes"):
                state = EHV_BLANK
            else:
                period_str = " ".join(get_value(row, 0).split()).lower()
                periods = []

                if (
                    period_str
                    == "monday to friday nov to feb (excluding "
                    + "22nd dec to 4th jan inclusive)"
                ):
                    periods.append(
                        {
                            "weekend": False,
                            "start-month": 11,
                            "start-day": 1,
                            "finish-month": 12,
                            "finish-day": 21,
                        }
                    )
                    periods.append(
                        {
                            "weekend": False,
                            "start-month": 1,
                            "start-day": 5,
                            "finish-month": 2,
                            "finish-day": "last",
                        }
                    )

                elif period_str in (
                    "monday to friday (including bank "
                    "holidays) november to february",
                    "monday to friday nov to feb",
                ):
                    periods.append(
                        {
                            "weekend": False,
                            "start-month": 11,
                            "start-day": 1,
                            "finish-month": 2,
                            "finish-day": "last",
                        }
                    )

                for slot in val_to_slots(get_value(row, 3)):
                    for period in periods:
                        bands.append(
                            {
                                **period,
                                "start_hour": slot["start"],
                                "finish_hour": slot["finish"],
                            }
                        )


def find_rates(file_name, file_like):
    rates = {"a_file_name": file_name}

    if not file_name.endswith(".xlsx"):
        raise BadRequest(f"The file extension for {file_name} isn't recognized.")

    book = openpyxl.load_workbook(file_like, data_only=True, read_only=True)

    for sheet in book.worksheets:
        title = sheet.title.strip().lower()
        if title.startswith("annex 1 "):
            tab_lv_hv(sheet, rates)
        elif title.startswith("annex 2 "):
            tab_ehv(sheet, rates)

    return rates
