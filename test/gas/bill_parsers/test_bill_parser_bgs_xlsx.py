from collections import defaultdict
from datetime import datetime as Datetime
from decimal import Decimal

from openpyxl import Workbook


from chellow.gas.bill_parser_bgs_xlsx import _make_raw_bills, _parse_row
from chellow.utils import ct_datetime, to_utc, utc_datetime


def test_parse_row():
    wb = Workbook()
    ws = wb.active

    rows = [
        [
            "Customer",
            "Account",
            "MPRN",
            "Bill",
            "SupplyAddress",
            "BillDate",
            "BillingPeriod",
            "kWhs",
            "ChargeType",
            "ChargePeriodFrom",
            "ChargePeriodEnd",
            "Quantity",
            "QUnit",
            "Charge",
            "CUnit",
            "Total",
        ],
        [
            "Zorin Industries",
            69389283,
            86762618,
            584773269,
            "Extinct Volcano",
            Datetime(2025, 6, 13),
            Datetime(2025, 5, 1),
            0,
            "Gas",
            Datetime(2025, 5, 1),
            Datetime(2025, 5, 31),
            0,
            "kWh",
            0.5311,
            "p/kWh",
            0.00,
        ],
    ]

    for row, row_values in enumerate(rows, start=1):
        for column, value in enumerate(row_values, start=1):
            ws.cell(row=row, column=column, value=value)

    row = 2
    title_row = tuple(ws.rows)[0]
    bills = {}
    _parse_row(bills, ws, row, title_row)


def test_make_raw_bills():
    wb = Workbook()
    ws = wb.active

    rows = [
        [
            "Customer",
            "Account",
            "MPRN",
            "Bill",
            "SupplyAddress",
            "BillDate",
            "BillingPeriod",
            "kWhs",
            "ChargeType",
            "ChargePeriodFrom",
            "ChargePeriodEnd",
            "Quantity",
            "QUnit",
            "Charge",
            "CUnit",
            "Total",
        ],
        [
            "Zorin Industries",
            69389283,
            86762618,
            584773269,
            "Extinct Volcano",
            Datetime(2025, 6, 13),
            Datetime(2025, 5, 1),
            0,
            "Gas",
            Datetime(2025, 5, 1),
            Datetime(2025, 5, 31),
            0,
            "kWh",
            0.5311,
            "p/kWh",
            0.00,
        ],
    ]

    for row, row_vals in enumerate(rows, start=1):
        for column, value in enumerate(row_vals, start=1):
            ws.cell(row=row, column=column, value=value)

    bills = _make_raw_bills(ws)
    expected_bills = [
        {
            "bill_type_code": "N",
            "mprn": "86762618",
            "reference": "584773269",
            "account": 69389283,
            "issue_date": utc_datetime(2025, 6, 12, 23, 0),
            "start_date": utc_datetime(2025, 4, 30, 23, 0),
            "finish_date": utc_datetime(2025, 5, 31, 22, 30),
            "kwh": Decimal("0"),
            "breakdown": {
                "admin_variable_gbp": Decimal("0.0"),
                "admin_variable_kwh": Decimal("0"),
                "admin_variable_rate": {Decimal("0.005311")},
            },
            "net_gbp": Decimal("0.00"),
            "vat_gbp": Decimal("0.00"),
            "gross_gbp": Decimal("0.00"),
            "raw_lines": [
                [
                    "Zorin Industries",
                    69389283,
                    86762618,
                    584773269,
                    "Extinct Volcano",
                    Datetime(2025, 6, 13, 0, 0),
                    Datetime(2025, 5, 1, 0, 0),
                    0,
                    "Gas",
                    Datetime(2025, 5, 1, 0, 0),
                    Datetime(2025, 5, 31, 0, 0),
                    0,
                    "kWh",
                    0.5311,
                    "p/kWh",
                    0.0,
                ]
            ],
            "reads": [],
        }
    ]
    assert bills == expected_bills


def test_make_raw_bills_part_month():
    wb = Workbook()
    ws = wb.active

    rows = [
        [
            "Customer",
            "Account",
            "MPRN",
            "Bill",
            "SupplyAddress",
            "BillDate",
            "BillingPeriod",
            "kWhs",
            "ChargeType",
            "ChargePeriodFrom",
            "ChargePeriodEnd",
            "Quantity",
            "QUnit",
            "Charge",
            "CUnit",
            "Total",
        ],
        [
            "Zorin Industries",
            69389283,
            86762618,
            584773269,
            "Extinct Volcano",
            Datetime(2025, 6, 13),
            Datetime(2025, 5, 1),
            0,
            "Gas",
            Datetime(2025, 5, 1),
            Datetime(2025, 5, 15),
            0,
            "kWh",
            0.5311,
            "p/kWh",
            0.00,
        ],
    ]

    for row, row_vals in enumerate(rows, start=1):
        for column, value in enumerate(row_vals, start=1):
            ws.cell(row=row, column=column, value=value)

    bills = _make_raw_bills(ws)
    expected_bills = [
        {
            "bill_type_code": "N",
            "mprn": "86762618",
            "reference": "584773269",
            "account": 69389283,
            "issue_date": utc_datetime(2025, 6, 12, 23, 0),
            "start_date": to_utc(ct_datetime(2025, 5, 1, 0, 0)),
            "finish_date": to_utc(ct_datetime(2025, 5, 15, 23, 30)),
            "kwh": Decimal("0"),
            "breakdown": defaultdict(
                int,
                {
                    "admin_variable_gbp": Decimal("0.0"),
                    "admin_variable_kwh": Decimal("0"),
                    "admin_variable_rate": {Decimal("0.005311")},
                },
            ),
            "net_gbp": Decimal("0.00"),
            "vat_gbp": Decimal("0.00"),
            "gross_gbp": Decimal("0.00"),
            "raw_lines": [
                [
                    "Zorin Industries",
                    69389283,
                    86762618,
                    584773269,
                    "Extinct Volcano",
                    Datetime(2025, 6, 13, 0, 0),
                    Datetime(2025, 5, 1, 0, 0),
                    0,
                    "Gas",
                    Datetime(2025, 5, 1, 0, 0),
                    Datetime(2025, 5, 15, 0, 0),
                    0,
                    "kWh",
                    0.5311,
                    "p/kWh",
                    0.0,
                ],
            ],
            "reads": [],
        }
    ]
    assert bills == expected_bills


def test_make_raw_bills_credit():
    wb = Workbook()
    ws = wb.active

    rows = [
        [
            "Customer",
            "Account",
            "MPRN",
            "Bill",
            "SupplyAddress",
            "BillDate",
            "BillingPeriod",
            "kWhs",
            "ChargeType",
            "ChargePeriodFrom",
            "ChargePeriodEnd",
            "Quantity",
            "QUnit",
            "Charge",
            "CUnit",
            "Total",
        ],
        [
            "Zorin Industries",
            69389283,
            86762618,
            "CR584773269",
            "Extinct Volcano",
            Datetime(2025, 6, 13),
            Datetime(2025, 5, 1),
            0,
            "Gas",
            Datetime(2025, 5, 1),
            Datetime(2025, 5, 31),
            10,
            "kWh",
            0.5311,
            "p/kWh",
            -5.00,
        ],
    ]

    for row, row_vals in enumerate(rows, start=1):
        for column, value in enumerate(row_vals, start=1):
            ws.cell(row=row, column=column, value=value)

    bills = _make_raw_bills(ws)
    expected_bills = [
        {
            "bill_type_code": "W",
            "mprn": "86762618",
            "reference": "CR584773269",
            "account": 69389283,
            "issue_date": utc_datetime(2025, 6, 12, 23, 0),
            "start_date": utc_datetime(2025, 4, 30, 23, 0),
            "finish_date": utc_datetime(2025, 5, 31, 22, 30),
            "kwh": Decimal("0"),
            "breakdown": {
                "admin_variable_gbp": Decimal("-5.00"),
                "admin_variable_kwh": Decimal("-10"),
                "admin_variable_rate": {Decimal("0.005311")},
            },
            "net_gbp": Decimal("0.00"),
            "vat_gbp": Decimal("0.00"),
            "gross_gbp": Decimal("0.00"),
            "raw_lines": [
                [
                    "Zorin Industries",
                    69389283,
                    86762618,
                    "CR584773269",
                    "Extinct Volcano",
                    Datetime(2025, 6, 13, 0, 0),
                    Datetime(2025, 5, 1, 0, 0),
                    0,
                    "Gas",
                    Datetime(2025, 5, 1, 0, 0),
                    Datetime(2025, 5, 31, 0, 0),
                    10,
                    "kWh",
                    0.5311,
                    "p/kWh",
                    -5.0,
                ]
            ],
            "reads": [],
        }
    ]
    assert bills == expected_bills


def test_make_raw_bills_multi_date():
    wb = Workbook()
    ws = wb.active

    rows = [
        [
            "Customer",
            "Account",
            "MPRN",
            "Bill",
            "SupplyAddress",
            "BillDate",
            "BillingPeriod",
            "kWhs",
            "ChargeType",
            "ChargePeriodFrom",
            "ChargePeriodEnd",
            "Quantity",
            "QUnit",
            "Charge",
            "CUnit",
            "Total",
        ],
        [
            "Zorin Industries",
            69389283,
            86762618,
            584773269,
            "Extinct Volcano",
            Datetime(2025, 6, 13),
            Datetime(2025, 5, 1),
            0,
            "Gas",
            Datetime(2025, 5, 1),
            Datetime(2025, 5, 15),
            0,
            "kWh",
            0.5311,
            "p/kWh",
            0.00,
        ],
        [
            "Zorin Industries",
            69389283,
            86762618,
            584773269,
            "Extinct Volcano",
            Datetime(2025, 6, 13),
            Datetime(2025, 5, 1),
            0,
            "Gas",
            Datetime(2025, 5, 16),
            Datetime(2025, 5, 31),
            0,
            "kWh",
            0.5311,
            "p/kWh",
            0.00,
        ],
    ]

    for row, row_vals in enumerate(rows, start=1):
        for column, value in enumerate(row_vals, start=1):
            ws.cell(row=row, column=column, value=value)

    bills = _make_raw_bills(ws)
    expected_bills = [
        {
            "bill_type_code": "N",
            "mprn": "86762618",
            "reference": "584773269",
            "account": 69389283,
            "issue_date": utc_datetime(2025, 6, 12, 23, 0),
            "start_date": to_utc(ct_datetime(2025, 5, 1, 0, 0)),
            "finish_date": to_utc(ct_datetime(2025, 5, 31, 23, 30)),
            "kwh": Decimal("0"),
            "breakdown": defaultdict(
                int,
                {
                    "admin_variable_gbp": Decimal("0.0"),
                    "admin_variable_kwh": Decimal("0"),
                    "admin_variable_rate": {Decimal("0.005311")},
                },
            ),
            "net_gbp": Decimal("0.00"),
            "vat_gbp": Decimal("0.00"),
            "gross_gbp": Decimal("0.00"),
            "raw_lines": [
                [
                    "Zorin Industries",
                    69389283,
                    86762618,
                    584773269,
                    "Extinct Volcano",
                    Datetime(2025, 6, 13, 0, 0),
                    Datetime(2025, 5, 1, 0, 0),
                    0,
                    "Gas",
                    Datetime(2025, 5, 1, 0, 0),
                    Datetime(2025, 5, 15, 0, 0),
                    0,
                    "kWh",
                    0.5311,
                    "p/kWh",
                    0.0,
                ],
                [
                    "Zorin Industries",
                    69389283,
                    86762618,
                    584773269,
                    "Extinct Volcano",
                    Datetime(2025, 6, 13, 0, 0),
                    Datetime(2025, 5, 1, 0, 0),
                    0,
                    "Gas",
                    Datetime(2025, 5, 16, 0, 0),
                    Datetime(2025, 5, 31, 0, 0),
                    0,
                    "kWh",
                    0.5311,
                    "p/kWh",
                    0.0,
                ],
            ],
            "reads": [],
        }
    ]
    assert bills == expected_bills


def test_make_raw_bills_VAT():
    wb = Workbook()
    ws = wb.active

    rows = [
        [
            "Customer",
            "Account",
            "MPRN",
            "Bill",
            "SupplyAddress",
            "BillDate",
            "BillingPeriod",
            "kWhs",
            "ChargeType",
            "ChargePeriodFrom",
            "ChargePeriodEnd",
            "Quantity",
            "QUnit",
            "Charge",
            "CUnit",
            "Total",
        ],
        [
            "Zorin Industries",
            69389283,
            86762618,
            584773269,
            "Extinct Volcano",
            Datetime(2025, 6, 13),
            Datetime(2025, 5, 1),
            0,
            "VAT",
            None,
            None,
            0,
            "kWh",
            0.5311,
            "p/kWh",
            0.00,
        ],
        [
            "Zorin Industries",
            69389283,
            86762618,
            584773269,
            "Extinct Volcano",
            Datetime(2025, 6, 13),
            Datetime(2025, 5, 1),
            0,
            "Gas",
            Datetime(2025, 5, 16),
            Datetime(2025, 5, 31),
            0,
            "kWh",
            0.5311,
            "p/kWh",
            0.00,
        ],
    ]

    for row, row_vals in enumerate(rows, start=1):
        for column, value in enumerate(row_vals, start=1):
            ws.cell(row=row, column=column, value=value)

    bills = _make_raw_bills(ws)
    expected_bills = [
        {
            "bill_type_code": "N",
            "mprn": "86762618",
            "reference": "584773269",
            "account": 69389283,
            "issue_date": utc_datetime(2025, 6, 12, 23, 0),
            "start_date": to_utc(ct_datetime(2025, 5, 16, 0, 0)),
            "finish_date": to_utc(ct_datetime(2025, 5, 31, 23, 30)),
            "kwh": Decimal("0"),
            "breakdown": defaultdict(
                int,
                {
                    "admin_variable_gbp": Decimal("0.0"),
                    "admin_variable_kwh": Decimal("0"),
                    "admin_variable_rate": {Decimal("0.005311")},
                    "vat_rate": {Decimal("0.005311")},
                },
            ),
            "net_gbp": Decimal("0.00"),
            "vat_gbp": Decimal("0.00"),
            "gross_gbp": Decimal("0.00"),
            "raw_lines": [
                [
                    "Zorin Industries",
                    69389283,
                    86762618,
                    584773269,
                    "Extinct Volcano",
                    Datetime(2025, 6, 13, 0, 0),
                    Datetime(2025, 5, 1, 0, 0),
                    0,
                    "VAT",
                    None,
                    None,
                    0,
                    "kWh",
                    0.5311,
                    "p/kWh",
                    0.0,
                ],
                [
                    "Zorin Industries",
                    69389283,
                    86762618,
                    584773269,
                    "Extinct Volcano",
                    Datetime(2025, 6, 13, 0, 0),
                    Datetime(2025, 5, 1, 0, 0),
                    0,
                    "Gas",
                    Datetime(2025, 5, 16, 0, 0),
                    Datetime(2025, 5, 31, 0, 0),
                    0,
                    "kWh",
                    0.5311,
                    "p/kWh",
                    0.0,
                ],
            ],
            "reads": [],
        }
    ]
    assert bills == expected_bills
