from datetime import datetime as Datetime
from decimal import Decimal

from openpyxl import Workbook

import pytest


from werkzeug.exceptions import BadRequest

from chellow.e.bill_parsers.engie_xlsx import _bd_add, _make_raw_bills, _parse_row
from chellow.utils import utc_datetime


def test_parse_row():
    wb = Workbook()
    ws = wb.active

    rows = [
        [
            "BillingEntity",
            "CustomerName",
            "CustomerNumber",
            "AccountName",
            "AccountNumber",
            "BillingAddresss",
            "BillNumber",
            "BillDate",
            "DueDate",
            "AcceptedDate",
            "BillPeriod",
            "AgreementNumber",
            "ProductBundle",
            "ProductItemName",
            "BillStatus",
            "ProductItemClass",
            "Type",
            "Description",
            "FromDate",
            "ToDate",
            "SalesTaxRate",
            "MeterPoint",
            "Usage",
            "UsageUnit",
            "Price",
            "Amount",
            "Currency",
            "Indicator",
            "RateName",
            "ProductName",
        ],
        [
            "Power Comany Ltd.",
            "Bill Paja",
            "556",
            "BILL PAJA",
            "883",
            "1 True Way",
            "672770",
            Datetime(2019, 3, 31, 23, 30),
            Datetime(2019, 3, 31, 23, 30),
            "",
            "2019-03-01 - 2019-03-31",
            "",
            "",
            "",
            "Draft",
            "",
            "Product",
            "Hand Held Read -",
            Datetime(2019, 3, 31, 23, 30),
            Datetime(2019, 3, 31, 0, 0),
            "",
            "2299999999929",
            "",
            "",
            "",
            "785",
            "GBP",
            "INV",
            "",
            "",
        ],
    ]

    for row, row_values in enumerate(rows, start=1):
        for column, value in enumerate(row_values, start=1):
            ws.cell(row=row, column=column, value=value)

    row = 2
    title_row = tuple(ws.rows)[0]
    bill = _parse_row(ws, row, title_row)
    assert bill["finish_date"] == utc_datetime(2019, 3, 31, 22, 30)


def test_bd_add():
    el_name = "duos_red"
    bd = {el_name: 0}
    val = None

    with pytest.raises(BadRequest):
        _bd_add(bd, el_name, val)


def test_bill_parser_engie_xls_billed_kwh():
    """
    Blank units should still give billed kWh
    """
    wb = Workbook()
    ws = wb.active

    rows = [
        [
            "BillingEntity",
            "CustomerName",
            "CustomerNumber",
            "AccountName",
            "AccountNumber",
            "BillingAddresss",
            "BillNumber",
            "BillDate",
            "DueDate",
            "AcceptedDate",
            "BillPeriod",
            "AgreementNumber",
            "ProductBundle",
            "ProductItemName",
            "BillStatus",
            "ProductItemClass",
            "Type",
            "Description",
            "FromDate",
            "ToDate",
            "SalesTaxRate",
            "MeterPoint",
            "Usage",
            "UsageUnit",
            "Price",
            "Amount",
            "Currency",
            "Indicator",
            "RateName",
            "ProductName",
        ],
        [
            "Mistral Wind Power Ltd.",
            "Bill Paja",
            "886572998",
            "Bill Paja",
            "869987122",
            "BA1 5TT",
            "99708221",
            Datetime(2019, 3, 31, 23, 30),
            42694,
            42629,
            "2016-08-01 - 2016-08-31",
            "458699",
            "Standard Deluxe",
            "Renewables Obligation (RO)",
            "Accepted",
            "Pass Thru - UK Electricity Cost Component",
            "Product",
            "Renewables Obligation (RO)",
            Datetime(2019, 3, 31, 23, 30),
            Datetime(2019, 3, 31, 23, 30),
            "Commercial UK Energy VAT",
            "2298132107763",
            27997.33,
            "",
            0.0971123676,
            "9224",
            "GBP",
            "INV",
            "Renewables Obligation (RO)",
            "",
        ],
    ]
    for row, row_vals in enumerate(rows, start=1):
        for column, value in enumerate(row_vals, start=1):
            ws.cell(row=row, column=column, value=value)

    row = 2
    title_row = tuple(ws.rows)[0]
    bill = _parse_row(ws, row, title_row)
    assert bill["kwh"] == Decimal("27997.33")


def test_make_raw_bills_vat():
    wb = Workbook()
    ws = wb.active

    rows = [
        [
            "BillingEntity",
            "CustomerName",
            "CustomerNumber",
            "AccountName",
            "AccountNumber",
            "BillingAddresss",
            "BillNumber",
            "BillDate",
            "DueDate",
            "AcceptedDate",
            "BillPeriod",
            "AgreementNumber",
            "ProductBundle",
            "ProductItemName",
            "BillStatus",
            "ProductItemClass",
            "Type",
            "Description",
            "FromDate",
            "ToDate",
            "SalesTaxRate",
            "MeterPoint",
            "Usage",
            "UsageUnit",
            "Price",
            "Amount",
            "Currency",
            "Indicator",
            "RateName",
            "ProductName",
        ],
        [
            "Mistral Wind Power Ltd.",
            "Bill Paja",
            "886572998",
            "Bill Paja",
            "869987122",
            "BA1 5TT",
            "99708221",
            Datetime(2019, 3, 31, 23, 30),
            Datetime(2019, 3, 31, 23, 30),
            Datetime(2019, 3, 31, 23, 30),
            "2016-08-01 - 2016-08-31",
            "",
            "",
            "",
            "Accepted",
            "",
            "Sales Tax",
            "Standard VAT@20%",
            "",
            "",
            "",
            "2298132107763",
            "",
            "",
            "",
            "9224",
            "GBP",
            "INV",
            "",
            "",
        ],
    ]

    for row, row_vals in enumerate(rows, start=1):
        for column, value in enumerate(row_vals, start=1):
            ws.cell(row=row, column=column, value=value)

    bills = _make_raw_bills(ws)
    expected_bills = [
        {
            "bill_type_code": "N",
            "kwh": Decimal("0"),
            "vat": Decimal("9224.00"),
            "net": Decimal("0.00"),
            "reads": [],
            "breakdown": {
                "vat_percentage": Decimal("20"),
            },
            "account": "22 9813 2107 763",
            "issue_date": utc_datetime(2019, 3, 31, 22, 30),
            "start_date": utc_datetime(2016, 7, 31, 23, 0),
            "finish_date": utc_datetime(2016, 8, 31, 22, 30),
            "mpan_core": "22 9813 2107 763",
            "reference": "99708221_2",
            "gross": Decimal("9224.00"),
        }
    ]
    assert bills == expected_bills
