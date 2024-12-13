from decimal import Decimal

from openpyxl import Workbook

from chellow.e.bill_parsers.edf_export_xlsx import _make_raw_bills
from chellow.utils import ct_datetime, to_utc


def test_make_raw_bills():
    wb = Workbook()
    statement_sheet = wb.worksheets[0]
    wb.create_sheet()  # HH Data
    wb.create_sheet()  # Calculations
    gduos_sheet = wb.create_sheet()  # GDUoS

    statement_sheet["D22"] = "01-May-2024"
    statement_sheet["F22"] = "31-May-2024"
    statement_sheet["H24"] = ct_datetime(2024, 8, 4)
    statement_sheet["D25"] = 1000000000003
    statement_sheet["D33"] = Decimal("302")
    statement_sheet["H33"] = Decimal("23.09")

    gduos_sheet["A2"] = 1000000000003
    gduos_sheet["B2"] = "2024-04"
    gduos_sheet["D2"] = "Capacity"
    gduos_sheet["E2"] = Decimal("41")
    gduos_sheet["F2"] = Decimal("4.97")
    gduos_sheet["G2"] = Decimal("83.12")

    bills = _make_raw_bills(wb)
    expected_bills = [
        {
            "account": "10 0000 0000 003",
            "bill_type_code": "N",
            "breakdown": {"ssp-gbp": Decimal("23.09"), "ssp-kwh": Decimal("302")},
            "finish_date": to_utc(ct_datetime(2024, 5, 31, 23, 30)),
            "gross": Decimal("23.09"),
            "issue_date": to_utc(ct_datetime(2024, 8, 4)),
            "kwh": Decimal("302"),
            "mpan_core": "10 0000 0000 003",
            "net": Decimal("23.09"),
            "reads": [],
            "reference": "20240501_20240531_20240804_10 0000 0000 003_ssp",
            "start_date": to_utc(ct_datetime(2024, 5, 1)),
            "vat": Decimal("0.00"),
        },
        {
            "account": "10 0000 0000 003",
            "bill_type_code": "N",
            "breakdown": {
                "duos-availability-gbp": Decimal("83.12"),
                "duos-availability-kva": Decimal("41"),
                "duos-availability-rate": [Decimal("4.97")],
            },
            "finish_date": to_utc(ct_datetime(2024, 4, 30, 23, 30)),
            "gross": Decimal("83.12"),
            "issue_date": to_utc(ct_datetime(2024, 8, 4)),
            "kwh": Decimal("0"),
            "mpan_core": "10 0000 0000 003",
            "net": Decimal("83.12"),
            "reads": [],
            "reference": "20240401_20240430_20240804_10 0000 0000 "
            "003_duos-availability",
            "start_date": to_utc(ct_datetime(2024, 4, 1)),
            "vat": Decimal("0.00"),
        },
    ]
    assert bills == expected_bills


def test_make_raw_bills_october():
    wb = Workbook()
    statement_sheet = wb.worksheets[0]
    wb.create_sheet()  # HH Data
    wb.create_sheet()  # Calculations
    gduos_sheet = wb.create_sheet()  # GDUoS

    statement_sheet["D22"] = "01-Oct-2024"
    statement_sheet["F22"] = "31-Oct-2024"
    statement_sheet["H24"] = ct_datetime(2024, 8, 4)
    statement_sheet["D25"] = 1000000000003
    statement_sheet["D33"] = Decimal("302")
    statement_sheet["H33"] = Decimal("23.09")

    gduos_sheet["A2"] = 1000000000003
    gduos_sheet["B2"] = "2024-10"
    gduos_sheet["D2"] = "Capacity"
    gduos_sheet["E2"] = Decimal("41")
    gduos_sheet["F2"] = Decimal("4.97")
    gduos_sheet["G2"] = Decimal("83.12")

    bills = _make_raw_bills(wb)
    expected_bills = [
        {
            "account": "10 0000 0000 003",
            "bill_type_code": "N",
            "breakdown": {"ssp-gbp": Decimal("23.09"), "ssp-kwh": Decimal("302")},
            "finish_date": to_utc(ct_datetime(2024, 10, 31, 23, 30)),
            "gross": Decimal("23.09"),
            "issue_date": to_utc(ct_datetime(2024, 8, 4)),
            "kwh": Decimal("302"),
            "mpan_core": "10 0000 0000 003",
            "net": Decimal("23.09"),
            "reads": [],
            "reference": "20241001_20241031_20240804_10 0000 0000 003_ssp",
            "start_date": to_utc(ct_datetime(2024, 10, 1)),
            "vat": Decimal("0.00"),
        },
        {
            "account": "10 0000 0000 003",
            "bill_type_code": "N",
            "breakdown": {
                "duos-availability-gbp": Decimal("83.12"),
                "duos-availability-kva": Decimal("41"),
                "duos-availability-rate": [Decimal("4.97")],
            },
            "finish_date": to_utc(ct_datetime(2024, 10, 31, 23, 30)),
            "gross": Decimal("83.12"),
            "issue_date": to_utc(ct_datetime(2024, 8, 4)),
            "kwh": Decimal("0"),
            "mpan_core": "10 0000 0000 003",
            "net": Decimal("83.12"),
            "reads": [],
            "reference": "20241001_20241031_20240804_10 0000 0000 "
            "003_duos-availability",
            "start_date": to_utc(ct_datetime(2024, 10, 1)),
            "vat": Decimal("0.00"),
        },
    ]
    assert bills == expected_bills
