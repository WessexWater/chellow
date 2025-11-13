from datetime import datetime as Datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook

from chellow.e.bill_parsers.nonsettlement_dc_stark_xlsx import Parser, _process_row
from chellow.utils import ct_datetime, to_utc, utc_datetime


def test_process_row(mocker, sess):
    rates = {
        "annual_rates": {
            "non_settlement": {"*": {"IP": {"*": {"gbp_per_meter": Decimal("12")}}}}
        }
    }
    mocker.patch(
        "chellow.e.bill_parsers.nonsettlement_dc_stark_xlsx.hh_rate", return_value=rates
    )
    caches = {}
    wb = Workbook()
    sheet = wb.worksheets[0]
    sheet.insert_rows(0, 2)
    sheet.insert_cols(0, 9)

    sheet["A1"] = "NAME"
    sheet["B1"] = "NMR"
    sheet["C1"] = "METER TYPE"
    sheet["D1"] = "IDENTIFIER"
    sheet["E1"] = "METER"
    sheet["F1"] = "MPAN REF"
    sheet["G1"] = "START"
    sheet["H1"] = "END"
    sheet["I1"] = "CHECK"

    sheet["A2"] = "Treglisson"
    sheet["B2"] = "Q997"
    sheet["C2"] = "VIS"
    sheet["D2"] = 101
    sheet["E2"] = "88jiuf ff"
    sheet["F2"] = 1472066139971
    sheet["G2"] = Datetime(2018, 6, 1)
    sheet["H2"] = Datetime(2018, 6, 30)
    sheet["I2"] = "Billed"
    bill = _process_row(caches, sess, sheet[1], sheet[2])
    assert bill == {
        "account": "14 7206 6139 971",
        "bill_type_code": "N",
        "breakdown": {
            "cop": [
                "5",
            ],
            "msn": [
                "88jiuf ff",
            ],
            "raw_lines": [],
            "vat": {
                20: {
                    "net": Decimal("1.00"),
                    "vat": Decimal("0.20"),
                },
            },
        },
        "elements": [
            {
                "breakdown": {
                    "comm": {
                        "IP",
                    },
                    "months": 1,
                    "rate": {
                        Decimal("12"),
                    },
                    "settlement-status": {
                        "non_settlement",
                    },
                },
                "finish_date": to_utc(ct_datetime(2018, 6, 30, 23, 30)),
                "name": "meter",
                "net": Decimal("1.00"),
                "start_date": to_utc(ct_datetime(2018, 6, 1, 0, 0)),
            },
        ],
        "finish_date": to_utc(ct_datetime(2018, 6, 30, 23, 30)),
        "gross": Decimal("1.20"),
        "issue_date": to_utc(ct_datetime(2018, 6, 1)),
        "kwh": Decimal("0"),
        "mpan_core": "14 7206 6139 971",
        "net": Decimal("1.00"),
        "reads": [],
        "reference": "20180601_20180630_20180601_14 7206 6139 971",
        "start_date": to_utc(ct_datetime(2018, 6, 1)),
        "vat": Decimal("0.20"),
    }


def test(mocker, sess):
    rates = {
        "annual_rates": {
            "non_settlement": {"*": {"IP": {"*": {"gbp_per_meter": Decimal("45")}}}}
        }
    }
    mocker.patch(
        "chellow.e.bill_parsers.nonsettlement_dc_stark_xlsx.hh_rate", return_value=rates
    )
    f = BytesIO()
    wb = Workbook()
    sheet = wb.worksheets[0]
    sheet.insert_rows(0, 2)
    sheet.insert_cols(0, 9)

    sheet["A1"] = "NAME"
    sheet["B1"] = "NMR"
    sheet["C1"] = "METER TYPE"
    sheet["D1"] = "IDENTIFIER"
    sheet["E1"] = "METER"
    sheet["F1"] = "MPAN REF"
    sheet["G1"] = "START"
    sheet["H1"] = "END"
    sheet["I1"] = "CHECK"

    sheet["A2"] = "Treglisson"
    sheet["B2"] = "Q997"
    sheet["C2"] = "VIS"
    sheet["D2"] = 101
    sheet["E2"] = "88jiuf ff"
    sheet["F2"] = 1472066139971
    sheet["G2"] = Datetime(2018, 6, 1)
    sheet["H2"] = Datetime(2018, 6, 30)
    sheet["I2"] = "Billed"

    wb.save(f)
    f.seek(0)
    parser = Parser(f)
    parser.make_raw_bills()


def test_old(mocker, sess):
    rates = {
        "annual_rates": {
            "non_settlement": {"*": {"IP": {"*": {"gbp_per_meter": Decimal("45")}}}}
        }
    }
    mocker.patch(
        "chellow.e.bill_parsers.nonsettlement_dc_stark_xlsx.hh_rate", return_value=rates
    )
    f = BytesIO()
    wb = Workbook()
    sheet = wb.worksheets[0]
    sheet.insert_rows(0, 2)
    sheet.insert_cols(0, 9)

    sheet["A1"] = "MPAN"
    sheet["B1"] = "WESSEX WATER SITE NAME"
    sheet["C1"] = "STARK IDENTIFIER"
    sheet["D1"] = "NMR"
    sheet["E1"] = "METER TYPE"
    sheet["F1"] = "IDENTIFIER"
    sheet["G1"] = "METER"
    sheet["H1"] = "MPAN REF"
    sheet["I1"] = "START"
    sheet["J1"] = "END"
    sheet["K1"] = "CHECK"

    sheet["A2"] = 1472066139971
    sheet["B2"] = ""
    sheet["C2"] = ""
    sheet["D2"] = "Q997"
    sheet["E2"] = "VIS"
    sheet["F2"] = 101
    sheet["G2"] = "88jiuf ff"
    sheet["H2"] = 1472066139971
    sheet["I2"] = Datetime(2018, 7, 1)
    sheet["J2"] = Datetime(2018, 7, 1)
    sheet["K2"] = "Billed"

    wb.save(f)
    f.seek(0)

    parser = Parser(f)
    parser.make_raw_bills()


def test_one_bill(mocker, sess):
    f = BytesIO()
    wb = Workbook()
    sheet = wb.worksheets[0]
    sheet.insert_rows(0, 2)
    sheet.insert_cols(0, 9)

    sheet["A1"] = "NAME"
    sheet["B1"] = "NMR"
    sheet["C1"] = "METER TYPE"
    sheet["D1"] = "IDENTIFIER"
    sheet["E1"] = "METER"
    sheet["F1"] = "MPAN REF"
    sheet["G1"] = "START"
    sheet["H1"] = "END"
    sheet["I1"] = "CHECK"

    sheet["A2"] = "Quantum Computer"
    sheet["B2"] = 65837766
    sheet["C2"] = "x600"
    sheet["D2"] = 7493477439
    sheet["E2"] = 7864739737
    sheet["F2"] = 2088757371777
    sheet["G2"] = Datetime(2022, 2, 1)
    sheet["H2"] = Datetime(2022, 2, 28)
    sheet["I2"] = "Billed"

    sheet["C6"].value = Datetime(2022, 3, 1)

    sheet["B12"].value = "2227651294270"
    sheet["C12"].value = "CS"
    sheet["D12"].value = "Settled"
    sheet["E12"].value = "Massive Field, Taunton, Somerset"
    sheet["F12"].value = Datetime(2022, 1, 1)
    sheet["G12"].value = Datetime(2022, 1, 31)
    sheet["H12"].value = "5.00"
    sheet["I12"].value = "0.12"
    sheet["J12"].value = "0.76"
    sheet["K12"].value = "0.99"

    wb.save(f)
    f.seek(0)
    rates = {
        "annual_rates": {
            "non_settlement": {"*": {"IP": {"*": {"gbp_per_meter": Decimal("45")}}}}
        }
    }
    mocker.patch(
        "chellow.e.bill_parsers.nonsettlement_dc_stark_xlsx.hh_rate", return_value=rates
    )
    p = Parser(f)
    result = p.make_raw_bills()

    expected = [
        {
            "bill_type_code": "N",
            "kwh": Decimal("0"),
            "vat": Decimal("0.75"),
            "net": Decimal("3.75"),
            "gross": Decimal("4.50"),
            "reads": [],
            "breakdown": {
                "raw_lines": [],
                "cop": ["5"],
                "msn": ["7864739737"],
                "vat": {
                    20: {
                        "net": Decimal("3.75"),
                        "vat": Decimal("0.75"),
                    },
                },
            },
            "account": "20 8875 7371 777",
            "issue_date": utc_datetime(2022, 2, 1),
            "start_date": utc_datetime(2022, 2, 1),
            "finish_date": utc_datetime(2022, 2, 28, 23, 30),
            "mpan_core": "20 8875 7371 777",
            "reference": "20220201_20220228_20220201_20 8875 7371 777",
            "elements": [
                {
                    "name": "meter",
                    "start_date": utc_datetime(2022, 2, 1),
                    "finish_date": utc_datetime(2022, 2, 28, 23, 30),
                    "net": Decimal("3.75"),
                    "breakdown": {
                        "rate": {Decimal("45")},
                        "months": 1,
                        "settlement-status": {"non_settlement"},
                        "comm": {"IP"},
                    },
                }
            ],
        }
    ]

    assert result == expected
