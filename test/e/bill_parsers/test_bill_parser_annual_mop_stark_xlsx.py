from datetime import datetime as Datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook

from chellow.e.bill_parsers.annual_mop_stark_xlsx import Parser
from chellow.utils import utc_datetime


def test_blank():
    f = BytesIO()
    wb = Workbook()
    wb.create_sheet("NHH")
    wb.save(f)
    f.seek(0)
    parser = Parser(f)
    assert parser.book.data_only


def test_no_bills(sess):
    f = BytesIO()
    wb = Workbook()
    wb.create_sheet("NHH")
    sheet = wb.worksheets[1]
    sheet.insert_rows(0, 10)
    sheet.insert_cols(0, 10)
    sheet["C6"].value = Datetime(2022, 3, 1)
    wb.save(f)
    f.seek(0)
    p = Parser(f)
    p.make_raw_bills()


def test_one_bill(sess):
    f = BytesIO()
    wb = Workbook()
    wb.create_sheet("NHH")
    sheet = wb.worksheets[1]
    sheet.insert_rows(0, 12)
    sheet.insert_cols(0, 11)
    sheet["C6"].value = Datetime(2022, 3, 1)

    sheet["B12"].value = "2227651294270"
    sheet["C12"].value = "CS"
    sheet["D12"].value = "Settled"
    sheet["E12"].value = "Massive Field, Taunton, Somerset"
    sheet["F12"].value = Datetime(2022, 1, 1)
    sheet["G12"].value = Datetime(2022, 1, 31)
    sheet["H12"].value = "5.00"
    sheet["I12"].value = "0.12"
    sheet["J12"].value = "0.76"
    sheet["K12"].value = "0.99"

    wb.save(f)
    f.seek(0)
    p = Parser(f)
    result = p.make_raw_bills()

    expected = [
        {
            "bill_type_code": "N",
            "kwh": Decimal("0"),
            "net": Decimal("0.12"),
            "vat": Decimal("0.76"),
            "gross": Decimal("0.99"),
            "reads": [],
            "breakdown": {
                "raw-lines": [],
                "comms": "CS",
                "settlement-status": ["settlement"],
                "meter-rate": [Decimal("5.00")],
                "meter-gbp": Decimal("0.12"),
            },
            "account": "22 2765 1294 270",
            "issue_date": utc_datetime(2022, 3, 1),
            "start_date": utc_datetime(2022, 1, 1),
            "finish_date": utc_datetime(2022, 1, 31, 23, 30),
            "mpan_core": "22 2765 1294 270",
            "reference": "20220101_20220131_20220301_22 2765 1294 270",
        }
    ]

    assert result == expected
