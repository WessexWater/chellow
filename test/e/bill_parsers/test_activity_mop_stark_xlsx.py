from datetime import datetime as Datetime
from io import BytesIO

from openpyxl import Workbook

from chellow.e.bill_parsers.activity_mop_stark_xlsx import Parser


def test_blank():
    f = BytesIO()
    wb = Workbook()
    wb.create_sheet("Summary")
    wb.create_sheet("NHH")
    wb.save(f)
    f.seek(0)
    parser = Parser(f)
    assert parser.book.data_only


def test_no_bills(sess):
    f = BytesIO()
    wb = Workbook()
    wb.create_sheet("NHH")
    sheet = wb.worksheets[1]
    sheet.insert_rows(0, 10)
    sheet.insert_cols(0, 10)
    sheet["B11"].value = "MPAN"
    sheet["C6"].value = Datetime(2022, 3, 1)
    wb.save(f)
    f.seek(0)
    p = Parser(f)
    p.make_raw_bills()
