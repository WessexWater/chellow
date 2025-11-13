from datetime import datetime as Datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook

from chellow.e.bill_parsers.settlement_dc_stark_xlsx import (
    Parser,
    _process_row,
    get_ct_date,
    get_gbp,
    make_column_lookup,
)
from chellow.utils import ct_datetime, to_ct, to_utc


def test_get_ct_date():
    dt_naive = Datetime(2024, 6, 30)

    wb = Workbook()
    sheet = wb.worksheets[0]
    sheet.insert_rows(0, 23)  # Have blank lines at the end
    sheet.insert_cols(0, 9)

    sheet["E12"] = dt_naive
    dt = get_ct_date(sheet, "E", 12)

    assert dt == to_ct(dt_naive)


def test_get_gbp():
    wb = Workbook()
    sheet = wb.worksheets[0]
    sheet.insert_rows(0, 23)  # Have blank lines at the end
    sheet.insert_cols(0, 9)

    sheet["E12"] = None
    gbp = get_gbp(sheet, "E", 12)

    assert gbp is None


def test_process_row(sess):

    wb = Workbook()
    sheet = wb.worksheets[0]
    sheet.insert_rows(0, 23)  # Have blank lines at the end
    sheet.insert_cols(0, 9)

    sheet["A6"] = "SSIL Billing Backing Data - Wessex"
    sheet["A7"] = "Date: 03/05/2024 09:16:58"

    sheet["G10"] = "Unmetered"
    sheet["J10"] = "Code 2"
    sheet["M10"] = "Code 3"
    sheet["P10"] = "Code 5"
    sheet["S10"] = "Code 10"
    sheet["V10"] = "GSM"
    sheet["Y10"] = "SavenergyOnline"
    sheet["AB10"] = "Meter Proving Tests"
    sheet["AE10"] = "Hand Held Visits (Ad hoc)"
    sheet["AH10"] = "Hand Held Visits (Regular)"
    sheet["AK10"] = "Annual Site Visits"

    sheet["A11"] = "Sage"
    sheet["B11"] = "MPAN"
    sheet["C11"] = "Site"
    sheet["D11"] = "Bill From"
    sheet["E11"] = "Bill To"
    sheet["F11"] = "Contract Reference"
    sheet["G11"] = "Unmetered Quant"
    sheet["H11"] = "Unmetered Rate"
    sheet["I11"] = "Quarterly Unmetered Charge"
    sheet["J11"] = "No. COP 2 Meters"
    sheet["K11"] = "Annual COP 2 DC/DA Rate"
    sheet["L11"] = "Quarterly COP 2 Charge"
    sheet["M11"] = "No. COP 3 Meters"
    sheet["N11"] = "Annual COP 3 DC/DA Rate"
    sheet["O11"] = "Quarterly COP 3 Charge"
    sheet["P11"] = "No. COP 5 Meters"
    sheet["Q11"] = "Annual COP 5 DC/DA Rate"
    sheet["R11"] = "Quarterly COP 5 Charge"
    sheet["S11"] = "No. COP 10 Meters"
    sheet["T11"] = "Annual COP 10 DC/DA Rate"
    sheet["U11"] = "Quarterly COP 10 Charge"
    sheet["V11"] = "No. GSM Meters"
    sheet["W11"] = "GSM Annual Rate"
    sheet["X11"] = "Quarterly GSM Charge"
    sheet["Y11"] = "No. Meters"
    sheet["Z11"] = "Annual SEO Rate"
    sheet["AA11"] = "Quarterly SEO Charge"
    sheet["AB11"] = "No. Meter Proving Test"
    sheet["AC11"] = "Meter Proving Test Rate"
    sheet["AD11"] = "Meter Proving Test Charge"
    sheet["AE11"] = "No. Hand Held Visits (Adhoc)"
    sheet["AF11"] = "Hand Held Visit (Adhoc) Rate"
    sheet["AG11"] = "Hand Held Visit (Adhoc) Charge"
    sheet["AH11"] = "No. Hand Held Visits (Regular)"
    sheet["AI11"] = "Hand Held Visit (Regular) Rate"
    sheet["AJ11"] = "Hand Held Visit (Regular) Charge"
    sheet["AK11"] = "No. Annual Site Visits"
    sheet["AL11"] = "Annual Site Visit Rate"
    sheet["AM11"] = "Annual Site Visit Charge"
    sheet["AN11"] = "Hand Held Visit Dates"
    sheet["AO11"] = "Grand Total"
    sheet["AP11"] = "VAT @ 20% "
    sheet["AQ11"] = "Grand Total"

    sheet["A12"] = "W006"
    sheet["B12"] = "2200013784589"
    sheet["C12"] = "NORTH PETHERTON STW, SOMERSET"
    sheet["D12"] = Datetime(2024, 4, 1)
    sheet["E12"] = Datetime(2024, 6, 30)
    sheet["F12"] = "GF002"
    sheet["G12"] = "0"
    sheet["H12"] = "0.00"
    sheet["I12"] = "0.00"
    sheet["J12"] = "0"
    sheet["K12"] = "0.00"
    sheet["L12"] = "0.00"
    sheet["M12"] = "0"
    sheet["N12"] = "99.50"
    sheet["O12"] = "0.00"
    sheet["P12"] = "1"
    sheet["Q12"] = "114.43"
    sheet["R12"] = "28.61"
    sheet["S12"] = "0"
    sheet["T12"] = "0.00"
    sheet["U12"] = "0.00"
    sheet["V12"] = "0"
    sheet["W12"] = "0.00"
    sheet["X12"] = "0.00"
    sheet["Y12"] = "1"
    sheet["Z12"] = "0.00"
    sheet["AA12"] = "0.00"
    sheet["AB12"] = "0"
    sheet["AC12"] = "25.00"
    sheet["AD12"] = "0.00"
    sheet["AE12"] = "0"
    sheet["AF12"] = "20.00"
    sheet["AG12"] = "0.00"
    sheet["AH12"] = "0"
    sheet["AI12"] = "0.00"
    sheet["AJ12"] = "0.00"
    sheet["AK12"] = "0"
    sheet["AL12"] = "20.00"
    sheet["AM12"] = "0.00"
    sheet["AN12"] = ""
    sheet["AO12"] = "28.61"
    sheet["AP12"] = "5.72"
    sheet["AQ12"] = "34.33"

    sheet["A13"] = ""  # Test copes with blank lines at the end

    row = 12
    issue_date = to_utc(ct_datetime(2025, 6, 2))
    column_lookup = make_column_lookup(sheet)
    bill = _process_row(sess, sheet, row, issue_date, column_lookup)

    assert bill == {
        "account": "22 0001 3784 589",
        "bill_type_code": "N",
        "breakdown": {
            "raw_lines": [],
            "settlement-status": [
                "settlement",
            ],
            "vat": {
                20: {
                    "net": Decimal("28.61"),
                    "vat": Decimal("5.72"),
                },
            },
        },
        "elements": [
            {
                "breakdown": {
                    "cop": {
                        "5",
                    },
                    "days": 91,
                    "mpan-days": 91,
                    "rate": {
                        Decimal("114.43"),
                    },
                },
                "finish_date": to_utc(ct_datetime(2024, 6, 30, 23, 30)),
                "name": "mpan",
                "net": Decimal("28.61"),
                "start_date": to_utc(ct_datetime(2024, 4, 1, 0, 0)),
            },
        ],
        "finish_date": to_utc(ct_datetime(2024, 6, 30, 23, 30)),
        "gross": Decimal("34.33"),
        "issue_date": issue_date,
        "kwh": Decimal("0"),
        "mpan_core": "22 0001 3784 589",
        "net": Decimal("28.61"),
        "reads": [],
        "reference": "20240401_20240630_20250602_22 0001 3784 589",
        "start_date": to_utc(ct_datetime(2024, 4, 1, 0, 0)),
        "vat": Decimal("5.72"),
    }


def test_process_row_element_sig_figs(sess):

    wb = Workbook()
    sheet = wb.worksheets[0]
    sheet.insert_rows(0, 23)  # Have blank lines at the end
    sheet.insert_cols(0, 9)

    sheet["A6"] = "SSIL Billing Backing Data - Wessex"
    sheet["A7"] = "Date: 03/05/2024 09:16:58"

    sheet["G10"] = "Unmetered"
    sheet["J10"] = "Code 2"
    sheet["M10"] = "Code 3"
    sheet["P10"] = "Code 5"
    sheet["S10"] = "Code 10"
    sheet["V10"] = "GSM"
    sheet["Y10"] = "SavenergyOnline"
    sheet["AB10"] = "Meter Proving Tests"
    sheet["AE10"] = "Hand Held Visits (Ad hoc)"
    sheet["AH10"] = "Hand Held Visits (Regular)"
    sheet["AK10"] = "Annual Site Visits"

    sheet["A11"] = "Sage"
    sheet["B11"] = "MPAN"
    sheet["C11"] = "Site"
    sheet["D11"] = "Bill From"
    sheet["E11"] = "Bill To"
    sheet["F11"] = "Contract Reference"
    sheet["G11"] = "Unmetered Quant"
    sheet["H11"] = "Unmetered Rate"
    sheet["I11"] = "Quarterly Unmetered Charge"
    sheet["J11"] = "No. COP 2 Meters"
    sheet["K11"] = "Annual COP 2 DC/DA Rate"
    sheet["L11"] = "Quarterly COP 2 Charge"
    sheet["M11"] = "No. COP 3 Meters"
    sheet["N11"] = "Annual COP 3 DC/DA Rate"
    sheet["O11"] = "Quarterly COP 3 Charge"
    sheet["P11"] = "No. COP 5 Meters"
    sheet["Q11"] = "Annual COP 5 DC/DA Rate"
    sheet["R11"] = "Quarterly COP 5 Charge"
    sheet["S11"] = "No. COP 10 Meters"
    sheet["T11"] = "Annual COP 10 DC/DA Rate"
    sheet["U11"] = "Quarterly COP 10 Charge"
    sheet["V11"] = "No. GSM Meters"
    sheet["W11"] = "GSM Annual Rate"
    sheet["X11"] = "Quarterly GSM Charge"
    sheet["Y11"] = "No. Meters"
    sheet["Z11"] = "Annual SEO Rate"
    sheet["AA11"] = "Quarterly SEO Charge"
    sheet["AB11"] = "No. Meter Proving Test"
    sheet["AC11"] = "Meter Proving Test Rate"
    sheet["AD11"] = "Meter Proving Test Charge"
    sheet["AE11"] = "No. Hand Held Visits (Adhoc)"
    sheet["AF11"] = "Hand Held Visit (Adhoc) Rate"
    sheet["AG11"] = "Hand Held Visit (Adhoc) Charge"
    sheet["AH11"] = "No. Hand Held Visits (Regular)"
    sheet["AI11"] = "Hand Held Visit (Regular) Rate"
    sheet["AJ11"] = "Hand Held Visit (Regular) Charge"
    sheet["AK11"] = "No. Annual Site Visits"
    sheet["AL11"] = "Annual Site Visit Rate"
    sheet["AM11"] = "Annual Site Visit Charge"
    sheet["AN11"] = "Hand Held Visit Dates"
    sheet["AO11"] = "Grand Total"
    sheet["AP11"] = "VAT @ 20% "
    sheet["AQ11"] = "Grand Total"

    sheet["A12"] = "W006"
    sheet["B12"] = "2200013784589"
    sheet["C12"] = "NORTH PETHERTON STW, SOMERSET"
    sheet["D12"] = Datetime(2024, 4, 1)
    sheet["E12"] = Datetime(2024, 6, 30)
    sheet["F12"] = "GF002"
    sheet["G12"] = "0"
    sheet["H12"] = "0.00"
    sheet["I12"] = "0.00"
    sheet["J12"] = "0"
    sheet["K12"] = "0.00"
    sheet["L12"] = "0.00"
    sheet["M12"] = "0"
    sheet["N12"] = "99.50"
    sheet["O12"] = "0.00"
    sheet["P12"] = "1"
    sheet["Q12"] = "114.43"
    sheet["R12"] = "28.6"
    sheet["S12"] = "0"
    sheet["T12"] = "0.00"
    sheet["U12"] = "0.00"
    sheet["V12"] = "0"
    sheet["W12"] = "0.00"
    sheet["X12"] = "0.00"
    sheet["Y12"] = "1"
    sheet["Z12"] = "0.00"
    sheet["AA12"] = "0.00"
    sheet["AB12"] = "0"
    sheet["AC12"] = "25.00"
    sheet["AD12"] = "0.00"
    sheet["AE12"] = "0"
    sheet["AF12"] = "20.00"
    sheet["AG12"] = "0.00"
    sheet["AH12"] = "0"
    sheet["AI12"] = "0.00"
    sheet["AJ12"] = "0.00"
    sheet["AK12"] = "0"
    sheet["AL12"] = "20.00"
    sheet["AM12"] = "0.00"
    sheet["AN12"] = ""
    sheet["AO12"] = "28.61"
    sheet["AP12"] = "5.72"
    sheet["AQ12"] = "34.33"

    sheet["A13"] = ""  # Test copes with blank lines at the end

    row = 12
    issue_date = to_utc(ct_datetime(2025, 6, 2))
    column_lookup = make_column_lookup(sheet)
    bill = _process_row(sess, sheet, row, issue_date, column_lookup)

    assert bill == {
        "account": "22 0001 3784 589",
        "bill_type_code": "N",
        "breakdown": {
            "raw_lines": [],
            "settlement-status": [
                "settlement",
            ],
            "vat": {
                20: {
                    "net": Decimal("28.61"),
                    "vat": Decimal("5.72"),
                },
            },
        },
        "elements": [
            {
                "breakdown": {
                    "cop": {
                        "5",
                    },
                    "days": 91,
                    "mpan-days": 91,
                    "rate": {
                        Decimal("114.43"),
                    },
                },
                "finish_date": to_utc(ct_datetime(2024, 6, 30, 23, 30)),
                "name": "mpan",
                "net": Decimal("28.60"),
                "start_date": to_utc(ct_datetime(2024, 4, 1, 0, 0)),
            },
            {
                "breakdown": {
                    "activity-name": {
                        "discrepancy",
                    },
                },
                "finish_date": to_utc(ct_datetime(2024, 6, 30, 23, 30)),
                "name": "activitiy",
                "net": Decimal("0.01"),
                "start_date": to_utc(ct_datetime(2024, 4, 1, 0, 0)),
            },
        ],
        "finish_date": to_utc(ct_datetime(2024, 6, 30, 23, 30)),
        "gross": Decimal("34.33"),
        "issue_date": issue_date,
        "kwh": Decimal("0"),
        "mpan_core": "22 0001 3784 589",
        "net": Decimal("28.61"),
        "reads": [],
        "reference": "20240401_20240630_20250602_22 0001 3784 589",
        "start_date": to_utc(ct_datetime(2024, 4, 1, 0, 0)),
        "vat": Decimal("5.72"),
    }
    assert bill["elements"][0]["net"].as_tuple().exponent == -2


def test_process_row_old_format(sess):

    wb = Workbook()
    sheet = wb.worksheets[0]
    sheet.insert_rows(0, 23)  # Have blank lines at the end
    sheet.insert_cols(0, 9)

    sheet["A6"] = "SSIL Billing Backing Data - Wessex"
    sheet["A7"] = "Date: 03/05/2024 09:16:58"

    sheet["G10"] = "Code 3"
    sheet["J10"] = "Code 5"
    sheet["M10"] = "GSM"
    sheet["P10"] = "SavenergyOnline"
    sheet["S10"] = "Meter Proving Tests"
    sheet["V10"] = "Hand Held Visits (Ad hoc)"
    sheet["Y10"] = "Hand Held Visits (Regular)"
    sheet["AB10"] = "Annual Site Visits"

    sheet["A11"] = "Sage"
    sheet["B11"] = "MPAN"
    sheet["C11"] = "Site"
    sheet["D11"] = "Bill From"
    sheet["E11"] = "Bill To"
    sheet["F11"] = "Contract Reference"
    sheet["G11"] = "No. COP 3 Meters"
    sheet["H11"] = "Annual COP 3 DC/DA Rate"
    sheet["I11"] = "COP 3 Charge"
    sheet["J11"] = "No. COP 5/10 Meters"
    sheet["K11"] = "Annual COP 5/10 DC/DA Rate"
    sheet["L11"] = "COP 5/10 Charge"
    sheet["M11"] = "No. GSM Meters"
    sheet["N11"] = "GSM Annual Rate"
    sheet["O11"] = "GSM Charge"
    sheet["P11"] = "No. Meters"
    sheet["Q11"] = "Annual SEO Rate"
    sheet["R11"] = "SEO Charge"
    sheet["S11"] = "No. Meter Proving Test"
    sheet["T11"] = "Meter Proving Test Rate"
    sheet["U11"] = "Meter Proving Test Charge"
    sheet["V11"] = "No. Hand Held Visits (Adhoc)"
    sheet["W11"] = "Hand Held Visit (Adhoc) Rate"
    sheet["X11"] = "Hand Held Visit (Adhoc) Charge"
    sheet["Y11"] = "No. Hand Held Visits (Regular)"
    sheet["Z11"] = "Hand Held Visit (Regular) Rate"
    sheet["AA11"] = "Hand Held Visit (Regular) Charge"
    sheet["AB11"] = "No. Annual Site Visits"
    sheet["AC11"] = "Annual Site Visit Rate"
    sheet["AD11"] = "Annual Site Visit Charge"
    sheet["AE11"] = "Hand Held Visit Dates"
    sheet["AF11"] = "Total Ex VAT"
    sheet["AG11"] = "VAT @ 20% "
    sheet["AH11"] = "Total In VAT"

    sheet["A12"] = "W006"
    sheet["B12"] = "2200013784589"
    sheet["C12"] = "NORTH PETHERTON STW, SOMERSET"
    sheet["D12"] = Datetime(2024, 4, 1)
    sheet["E12"] = Datetime(2024, 6, 30)
    sheet["F12"] = "GF002"
    sheet["G12"] = "0"
    sheet["H12"] = "99.50"
    sheet["I12"] = "0.00"
    sheet["J12"] = "1"
    sheet["K12"] = "114.43"
    sheet["L12"] = "28.61"
    sheet["M12"] = "0"
    sheet["N12"] = "0.00"
    sheet["O12"] = "0.00"
    sheet["P12"] = "1"
    sheet["Q12"] = "0.00"
    sheet["R12"] = "0.00"
    sheet["S12"] = "0"
    sheet["T12"] = "25.00"
    sheet["U12"] = "0.00"
    sheet["V12"] = "0"
    sheet["W12"] = "20.00"
    sheet["X12"] = "0.00"
    sheet["Y12"] = "0"
    sheet["Z12"] = "0.00"
    sheet["AA12"] = "0.00"
    sheet["AB12"] = "0"
    sheet["AC12"] = "20.00"
    sheet["AD12"] = "0.00"
    sheet["AE12"] = ""
    sheet["AF12"] = "28.61"
    sheet["AG12"] = "5.72"
    sheet["AH12"] = "34.33"

    sheet["A13"] = ""  # Test copes with blank lines at the end

    row = 12
    issue_date = to_utc(ct_datetime(2025, 6, 2))
    column_lookup = make_column_lookup(sheet)
    bill = _process_row(sess, sheet, row, issue_date, column_lookup)

    assert bill == {
        "account": "22 0001 3784 589",
        "bill_type_code": "N",
        "breakdown": {
            "raw_lines": [],
            "settlement-status": [
                "settlement",
            ],
            "vat": {
                20: {
                    "net": Decimal("28.61"),
                    "vat": Decimal("5.72"),
                },
            },
        },
        "elements": [
            {
                "breakdown": {
                    "cop": {
                        "5",
                    },
                    "days": 91,
                    "mpan-days": 91,
                    "rate": {
                        Decimal("114.43"),
                    },
                },
                "finish_date": to_utc(ct_datetime(2024, 6, 30, 23, 30)),
                "name": "mpan",
                "net": Decimal("28.61"),
                "start_date": to_utc(ct_datetime(2024, 4, 1, 0, 0)),
            },
        ],
        "finish_date": to_utc(ct_datetime(2024, 6, 30, 23, 30)),
        "gross": Decimal("34.33"),
        "issue_date": issue_date,
        "kwh": Decimal("0"),
        "mpan_core": "22 0001 3784 589",
        "net": Decimal("28.61"),
        "reads": [],
        "reference": "20240401_20240630_20250602_22 0001 3784 589",
        "start_date": to_utc(ct_datetime(2024, 4, 1, 0, 0)),
        "vat": Decimal("5.72"),
    }


def test_make_raw_bills(sess):

    f = BytesIO()
    wb = Workbook()
    sheet = wb.worksheets[0]
    sheet.insert_rows(0, 23)  # Have blank lines at the end
    sheet.insert_cols(0, 9)

    sheet["A6"] = "SSIL Billing Backing Data - Wessex"
    sheet["A7"] = "Date: 03/05/2024 09:16:58"

    sheet["G10"] = "Unmetered"
    sheet["J10"] = "Code 2"
    sheet["M10"] = "Code 3"
    sheet["P10"] = "Code 5"
    sheet["S10"] = "Code 10"
    sheet["V10"] = "GSM"
    sheet["Y10"] = "SavenergyOnline"
    sheet["AB10"] = "Meter Proving Tests"
    sheet["AE10"] = "Hand Held Visits (Ad hoc)"
    sheet["AH10"] = "Hand Held Visits (Regular)"
    sheet["AK10"] = "Annual Site Visits"

    sheet["A11"] = "Sage"
    sheet["B11"] = "MPAN"
    sheet["C11"] = "Site"
    sheet["D11"] = "Bill From"
    sheet["E11"] = "Bill To"
    sheet["F11"] = "Contract Reference"
    sheet["G11"] = "Unmetered Quant"
    sheet["H11"] = "Unmetered Rate"
    sheet["I11"] = "Quarterly Unmetered Charge"
    sheet["J11"] = "No. COP 2 Meters"
    sheet["K11"] = "Annual COP 2 DC/DA Rate"
    sheet["L11"] = "Quarterly COP 2 Charge"
    sheet["M11"] = "No. COP 3 Meters"
    sheet["N11"] = "Annual COP 3 DC/DA Rate"
    sheet["O11"] = "Quarterly COP 3 Charge"
    sheet["P11"] = "No. COP 5 Meters"
    sheet["Q11"] = "Annual COP 5 DC/DA Rate"
    sheet["R11"] = "Quarterly COP 5 Charge"
    sheet["S11"] = "No. COP 10 Meters"
    sheet["T11"] = "Annual COP 10 DC/DA Rate"
    sheet["U11"] = "Quarterly COP 10 Charge"
    sheet["V11"] = "No. GSM Meters"
    sheet["W11"] = "GSM Annual Rate"
    sheet["X11"] = "Quarterly GSM Charge"
    sheet["Y11"] = "No. Meters"
    sheet["Z11"] = "Annual SEO Rate"
    sheet["AA11"] = "Quarterly SEO Charge"
    sheet["AB11"] = "No. Meter Proving Test"
    sheet["AC11"] = "Meter Proving Test Rate"
    sheet["AD11"] = "Meter Proving Test Charge"
    sheet["AE11"] = "No. Hand Held Visits (Adhoc)"
    sheet["AF11"] = "Hand Held Visit (Adhoc) Rate"
    sheet["AG11"] = "Hand Held Visit (Adhoc) Charge"
    sheet["AH11"] = "No. Hand Held Visits (Regular)"
    sheet["AI11"] = "Hand Held Visit (Regular) Rate"
    sheet["AJ11"] = "Hand Held Visit (Regular) Charge"
    sheet["AK11"] = "No. Annual Site Visits"
    sheet["AL11"] = "Annual Site Visit Rate"
    sheet["AM11"] = "Annual Site Visit Charge"
    sheet["AN11"] = "Hand Held Visit Dates"
    sheet["AO11"] = "Grand Total"
    sheet["AP11"] = "VAT @ 20% "
    sheet["AQ11"] = "Grand Total"

    sheet["A12"] = "W006"
    sheet["B12"] = "2200013784589"
    sheet["C12"] = (
        "NORTH PETHERTON STW, PARKERS FIELD, NORTH PETHERTON, BRIDGWATER, SOMERSET"
    )
    sheet["D12"] = Datetime(2024, 4, 1)
    sheet["E12"] = Datetime(2024, 6, 30)
    sheet["F12"] = "GF002"
    sheet["G12"] = "0"
    sheet["H12"] = "0.00"
    sheet["I12"] = "0.00"
    sheet["J12"] = "0"
    sheet["K12"] = "0.00"
    sheet["L12"] = "0.00"
    sheet["M12"] = "0"
    sheet["N12"] = "99.50"
    sheet["O12"] = "0.00"
    sheet["P12"] = "1"
    sheet["Q12"] = "114.43"
    sheet["R12"] = "28.61"
    sheet["S12"] = "0"
    sheet["T12"] = "0.00"
    sheet["U12"] = "0.00"
    sheet["V12"] = "0"
    sheet["W12"] = "0.00"
    sheet["X12"] = "0.00"
    sheet["Y12"] = "1"
    sheet["Z12"] = "0.00"
    sheet["AA12"] = "0.00"
    sheet["AB12"] = "0"
    sheet["AC12"] = "25.00"
    sheet["AD12"] = "0.00"
    sheet["AE12"] = "0"
    sheet["AF12"] = "20.00"
    sheet["AG12"] = "0.00"
    sheet["AH12"] = "0"
    sheet["AI12"] = "0.00"
    sheet["AJ12"] = "0.00"
    sheet["AK12"] = "0"
    sheet["AL12"] = "20.00"
    sheet["AM12"] = "0.00"
    sheet["AN12"] = ""
    sheet["AO12"] = "28.61"
    sheet["AP12"] = "5.72"
    sheet["AQ12"] = "34.33"

    sheet["A13"] = "W006"
    sheet["B13"] = "1470001573345"
    sheet["C13"] = (
        "WICKWAR SEWAGE WORKS, OFF STATION ROAD, WICKWAR, WOTTON-UNDER-EDGE, "
        "GLOUCESTERSHIRE"
    )
    sheet["D13"] = Datetime(2024, 6, 12)
    sheet["E13"] = Datetime(2024, 6, 30)
    sheet["F13"] = "GF002"
    sheet["G13"] = "0"
    sheet["H13"] = "0.00"
    sheet["I13"] = "0.00"
    sheet["J13"] = "0"
    sheet["K13"] = "0.00"
    sheet["L13"] = "0.00"
    sheet["M13"] = "0"
    sheet["N13"] = "99.50"
    sheet["O13"] = "0.00"
    sheet["P13"] = "1"
    sheet["Q13"] = "114.43"
    sheet["R13"] = "5.96"
    sheet["S13"] = "0"
    sheet["T13"] = "0.00"
    sheet["U13"] = "0.00"
    sheet["V13"] = "0"
    sheet["W13"] = "0.00"
    sheet["X13"] = "0.00"
    sheet["Y13"] = "0.6333"
    sheet["Z13"] = "0.00"
    sheet["AA13"] = "0.00"
    sheet["AB13"] = "0"
    sheet["AC13"] = "25.00"
    sheet["AD13"] = "0.00"
    sheet["AE13"] = "0"
    sheet["AF13"] = "20.00"
    sheet["AG13"] = "0.00"
    sheet["AH13"] = "0"
    sheet["AI13"] = "0.00"
    sheet["AJ13"] = "0.00"
    sheet["AK13"] = "0"
    sheet["AL13"] = "20.00"
    sheet["AM13"] = "0.00"
    sheet["AN13"] = ""
    sheet["AO13"] = "5.96"
    sheet["AP13"] = "1.19"
    sheet["AQ13"] = "7.15"

    sheet["A14"] = ""  # Test copes with blank lines at the end

    wb.save(f)
    f.seek(0)
    p = Parser(f)
    raw_bills = p.make_raw_bills()
    assert raw_bills == [
        {
            "account": "22 0001 3784 589",
            "bill_type_code": "N",
            "breakdown": {
                "raw_lines": [],
                "settlement-status": [
                    "settlement",
                ],
                "vat": {
                    20: {
                        "net": Decimal("28.61"),
                        "vat": Decimal("5.72"),
                    },
                },
            },
            "elements": [
                {
                    "breakdown": {
                        "cop": {
                            "5",
                        },
                        "days": 91,
                        "mpan-days": 91,
                        "rate": {
                            Decimal("114.43"),
                        },
                    },
                    "finish_date": to_utc(ct_datetime(2024, 6, 30, 23, 30)),
                    "name": "mpan",
                    "net": Decimal("28.61"),
                    "start_date": to_utc(ct_datetime(2024, 4, 1, 0, 0)),
                },
            ],
            "finish_date": to_utc(ct_datetime(2024, 6, 30, 23, 30)),
            "gross": Decimal("34.33"),
            "issue_date": to_utc(ct_datetime(2024, 5, 3, 9, 16)),
            "kwh": Decimal("0"),
            "mpan_core": "22 0001 3784 589",
            "net": Decimal("28.61"),
            "reads": [],
            "reference": "20240401_20240630_20240503_22 0001 3784 589",
            "start_date": to_utc(ct_datetime(2024, 4, 1, 0, 0)),
            "vat": Decimal("5.72"),
        },
        {
            "account": "14 7000 1573 345",
            "bill_type_code": "N",
            "breakdown": {
                "raw_lines": [],
                "settlement-status": [
                    "settlement",
                ],
                "vat": {
                    20: {
                        "net": Decimal("5.96"),
                        "vat": Decimal("1.19"),
                    },
                },
            },
            "elements": [
                {
                    "breakdown": {
                        "cop": {
                            "5",
                        },
                        "days": 19,
                        "mpan-days": 19,
                        "rate": {
                            Decimal("114.43"),
                        },
                    },
                    "finish_date": to_utc(ct_datetime(2024, 6, 30, 23, 30)),
                    "name": "mpan",
                    "net": Decimal("5.96"),
                    "start_date": to_utc(ct_datetime(2024, 6, 12, 0, 0)),
                },
            ],
            "finish_date": to_utc(ct_datetime(2024, 6, 30, 23, 30)),
            "gross": Decimal("7.15"),
            "issue_date": to_utc(ct_datetime(2024, 5, 3, 9, 16)),
            "kwh": Decimal("0"),
            "mpan_core": "14 7000 1573 345",
            "net": Decimal("5.96"),
            "reads": [],
            "reference": "20240612_20240630_20240503_14 7000 1573 345",
            "start_date": to_utc(ct_datetime(2024, 6, 12, 0, 0)),
            "vat": Decimal("1.19"),
        },
    ]
