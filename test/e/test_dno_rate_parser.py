from decimal import Decimal
from io import BytesIO
from zipfile import ZipFile

from openpyxl import Workbook

import pytest

from chellow.e.dno_rate_parser import (
    find_rates,
    str_to_hr,
    tab_ehv,
    tab_llfs,
    tab_lv_hv,
    to_llfcs,
    to_pcs,
)


@pytest.mark.parametrize(
    "llfc_str,llfc_list",
    [
        ["H00-H01", ["H00", "H01"]],
        ["310 - 311 inclusive", ["310", "311"]],
        [
            """310
311
""",
            ["310", "311"],
        ],
    ],
)
def test_to_llfcs(mocker, llfc_str, llfc_list):
    actual = to_llfcs(llfc_str)
    assert actual == llfc_list


@pytest.mark.parametrize(
    "pc_str,pc_list", [["3 to 8 or 0", ["03", "04", "05", "06", "07", "08", "00"]]]
)
def test_to_pcs(mocker, pc_str, pc_list):
    mocker.patch("chellow.e.dno_rate_parser.get_value", return_value=pc_str)

    row = mocker.Mock()
    idx = mocker.Mock()
    actual = to_pcs(row, idx)
    assert actual == pc_list


@pytest.mark.parametrize("hr_str,hr", [["1600", Decimal("16")]])
def test_str_to_hr(mocker, hr_str, hr):
    actual = str_to_hr(hr_str)
    assert actual == hr


def test_find_rates():
    file_name = "a.zip"
    file_like = BytesIO()
    with ZipFile(file_like, "w") as zf:
        zf.writestr("b.txt", "")
    file_like.seek(0)
    find_rates(file_name, file_like)


def test_tab_llfs():
    wb = Workbook()
    wb.create_sheet("Annex 5 LLFs")
    sheet = wb.worksheets[1]
    sheet.insert_rows(0, 23)
    sheet.insert_cols(0, 9)
    sheet["A14"].value = "Metered voltage"
    sheet["A15"].value = "Low-voltage network"
    sheet["H15"].value = "100,ABB"

    vls = tab_llfs(sheet)
    assert vls == [
        {"code": "100", "voltage_level": "LV", "is_substation": False},
        {"code": "ABB", "voltage_level": "LV", "is_substation": False},
    ]


def test_tab_llfs_blank():
    wb = Workbook()
    wb.create_sheet("Annex 5 LLFs")
    sheet = wb.worksheets[1]
    sheet.insert_rows(0, 23)
    sheet.insert_cols(0, 9)
    sheet["A14"].value = "Metered voltage"
    sheet["A15"].value = "Low-voltage network"
    sheet["H15"].value = None

    vls = tab_llfs(sheet)
    assert vls == []


def test_tab_lv_hv():
    wb = Workbook()
    wb.create_sheet("Annex 1 LV, HV and UMS charges")
    sheet = wb.worksheets[1]
    sheet.insert_rows(0, 23)
    sheet.insert_cols(0, 12)
    sheet["A5"].value = "Time periods"
    sheet["A6"].value = "Saturday and Sunday\nAll year"
    sheet["B6"].value = ""
    sheet["C6"].value = "09:30 - 21:30"
    sheet["E6"].value = "00:00 - 09:30\n21:20 - 24:00"
    sheet["A7"].value = "Notes"

    sheet["G6"].value = (
        "Monday to Friday\n(Including Bank Holidays)\n Nov to Feb Inclusive "
        "(excluding 22nd Dec to 4th Jan inclusive)"
    )
    sheet["H6"].value = ""
    sheet["I6"].value = "17:00 to 19:00"
    sheet["J6"].value = "07:30 to 17:00\n19:00 to 21:30"
    sheet["G7"].value = "Notes"

    rates = {}
    tab_lv_hv(sheet, rates)
    assert rates == {
        "bands": [
            {
                "band": "amber",
                "finish": Decimal("21.5"),
                "start": Decimal("9.5"),
                "weekend": True,
            }
        ],
        "ums_bands": [
            {
                "band": "black",
                "finish-month": 12,
                "finish-day": 21,
                "finish-decimal-hour": Decimal("19"),
                "start-month": 11,
                "start-day": 1,
                "start-decimal-hour": Decimal("17"),
                "weekend": False,
            },
            {
                "band": "black",
                "finish-day": "last",
                "finish-decimal-hour": Decimal("19"),
                "finish-month": 2,
                "start-day": 4,
                "start-decimal-hour": Decimal("17"),
                "start-month": 1,
                "weekend": False,
            },
            {
                "band": "yellow",
                "finish-day": 21,
                "finish-decimal-hour": Decimal("17"),
                "finish-month": 12,
                "start-day": 1,
                "start-decimal-hour": Decimal("7.5"),
                "start-month": 11,
                "weekend": False,
            },
            {
                "band": "yellow",
                "finish-day": "last",
                "finish-decimal-hour": Decimal("17"),
                "finish-month": 2,
                "start-day": 4,
                "start-decimal-hour": Decimal("7.5"),
                "start-month": 1,
                "weekend": False,
            },
            {
                "band": "yellow",
                "finish-day": 21,
                "finish-decimal-hour": Decimal("21.5"),
                "finish-month": 12,
                "start-day": 1,
                "start-decimal-hour": Decimal("19"),
                "start-month": 11,
                "weekend": False,
            },
            {
                "band": "yellow",
                "finish-day": "last",
                "finish-decimal-hour": Decimal("21.5"),
                "finish-month": 2,
                "start-day": 4,
                "start-decimal-hour": Decimal("19"),
                "start-month": 1,
                "weekend": False,
            },
        ],
        "tariffs": {},
    }


def test_tab_ehv():
    wb = Workbook()
    wb.create_sheet("Annex 2 EHV charges")
    sheet = wb.worksheets[1]
    sheet.insert_rows(0, 10)
    sheet.insert_cols(0, 15)
    sheet["A5"].value = "time periods"
    sheet["A6"].value = (
        "Monday to Friday (Including Bank Holidays) Nov to Feb Inclusive (excluding "
        "22nd Dec to 4th Jan inclusive)"
    )
    sheet["D6"].value = "17:00 to 19:00"
    sheet["A7"].value = "Notes"
    sheet["A10"].value = "Import Unique Identifier"
    sheet["B10"].value = "LLFC"
    sheet["E10"].value = "LLFC"
    sheet["H10"].value = "Residual Charging Band"
    sheet["I10"].value = "Import Super Red unit charge (p/kWh)"
    sheet["J10"].value = "Import fixed charge (p/day)"
    sheet["K10"].value = "Import capacity charge (p/kVA/day)"
    sheet["L10"].value = "Import exceeded capacity charge (p/kVA/day)"
    sheet["B11"].value = "198"
    sheet["H11"].value = "2"
    sheet["I11"].value = "4.124"
    sheet["J11"].value = "18767.69"
    sheet["K11"].value = "1.40"
    sheet["L11"].value = "1.40"

    rates = {}
    tab_ehv(sheet, rates)
    assert rates == {
        "tariffs": {
            "198": {
                "super-red-gbp-per-kwh": Decimal("0.04124"),
                "gbp-per-mpan-per-day": Decimal("187.67690"),
                "gbp-per-kva-per-day": Decimal("0.01400"),
                "excess-gbp-per-kva-per-day": Decimal("0.01400"),
                "description": "Designated EHV2",
            }
        },
        "super_red": [
            {
                "finish-day": 21,
                "finish-decimal-hour": Decimal("19"),
                "finish-month": 12,
                "start-day": 1,
                "start-decimal-hour": Decimal("17"),
                "start-month": 11,
                "weekend": False,
            },
            {
                "finish-day": "last",
                "finish-decimal-hour": Decimal("19"),
                "finish-month": 2,
                "start-day": 4,
                "start-decimal-hour": Decimal("17"),
                "start-month": 1,
                "weekend": False,
            },
        ],
    }


def test_tab_ehv_null_band():
    wb = Workbook()
    wb.create_sheet("Annex 2 EHV charges")
    sheet = wb.worksheets[1]
    sheet.insert_rows(0, 10)
    sheet.insert_cols(0, 15)
    sheet["A10"].value = "Import Unique Identifier"
    sheet["B10"].value = "LLFC"
    sheet["E10"].value = "LLFC"
    sheet["H10"].value = "Residual Charging Band"
    sheet["I10"].value = "Import Super Red unit charge (p/kWh)"
    sheet["J10"].value = "Import fixed charge (p/day)"
    sheet["K10"].value = "Import capacity charge (p/kVA/day)"
    sheet["L10"].value = "Import exceeded capacity charge (p/kVA/day)"
    sheet["B11"].value = "198"
    sheet["H11"].value = None
    sheet["I11"].value = "4.124"
    sheet["J11"].value = "18767.69"
    sheet["K11"].value = "1.40"
    sheet["L11"].value = "1.40"

    rates = {}
    tab_ehv(sheet, rates)
    assert rates == {
        "tariffs": {
            "198": {
                "super-red-gbp-per-kwh": Decimal("0.04124"),
                "gbp-per-mpan-per-day": Decimal("187.67690"),
                "gbp-per-kva-per-day": Decimal("0.01400"),
                "excess-gbp-per-kva-per-day": Decimal("0.01400"),
                "description": "Designated EHV0",
            }
        },
        "super_red": [],
    }
